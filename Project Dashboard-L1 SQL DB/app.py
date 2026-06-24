import io
import json
import os
import re
import sqlite3
import traceback

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template_string, send_file

load_dotenv()

from db_utils import DB_PATH

app = Flask(__name__)

# --- PVT Project Dashboard route ---
@app.route('/pvt')
def pvt_dashboard():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Filter for PVT project data (assumes project_name or similar field distinguishes)
    cursor.execute("SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data, project_name FROM dashboard_data_pvt ORDER BY id DESC")
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        tc_summary = json.loads(rec[2])
        open_defects = json.loads(rec[3])
        detailed_tc_headers = json.loads(rec[4])
        detailed_tc_data = json.loads(rec[5])
        project_name = rec[6] if len(rec) > 6 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                # Records are ordered newest-first; keep the first seen for each (PID, Project Name)
                if key not in dedup_dict:
                    dedup_dict[key] = {
                        "cells": cells,
                        "tc_summary": tc_summary,
                        "open_defects": open_defects,
                        "detailed_tc_headers": detailed_tc_headers,
                        "detailed_tc_data": detailed_tc_data
                    }
            except Exception:
                all_rows.append({
                    "cells": cells,
                    "tc_summary": tc_summary,
                    "open_defects": open_defects,
                    "detailed_tc_headers": detailed_tc_headers,
                    "detailed_tc_data": detailed_tc_data
                })
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='PVT Project Dashboard - L1 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )


# --- ETE Billing Dashboard route ---
@app.route('/ete-billing')
def ete_billing_dashboard():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Filter for ETE Billing project data (assumes project_name or similar field distinguishes)
    cursor.execute("SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data, project_name FROM dashboard_data_billing ORDER BY id DESC")
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        tc_summary = json.loads(rec[2])
        open_defects = json.loads(rec[3])
        detailed_tc_headers = json.loads(rec[4])
        detailed_tc_data = json.loads(rec[5])
        project_name = rec[6] if len(rec) > 6 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                # Records are ordered newest-first; keep the first seen for each (PID, Project Name)
                if key not in dedup_dict:
                    dedup_dict[key] = {
                        "cells": cells,
                        "tc_summary": tc_summary,
                        "open_defects": open_defects,
                        "detailed_tc_headers": detailed_tc_headers,
                        "detailed_tc_data": detailed_tc_data
                    }
            except Exception:
                all_rows.append({
                    "cells": cells,
                    "tc_summary": tc_summary,
                    "open_defects": open_defects,
                    "detailed_tc_headers": detailed_tc_headers,
                    "detailed_tc_data": detailed_tc_data
                })
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='ETE Billing Project Dashboard - L1 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )
dashboard_html = r'''
<!DOCTYPE html>
<html lang="en">
<head>
    <!-- Font Awesome for Home Icon -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css">
    <!-- Google Fonts: Poppins -->
    <link href="https://fonts.googleapis.com/css?family=Poppins:600,700&display=swap" rel="stylesheet">
    <meta charset="UTF-8">
    <title>Project Dashboard</title>
    <style>
        /* Reduce font size for Project ID column */
        #mainTable td:first-child {
                font-size: 11px !important;
                text-align: center !important;
                vertical-align: middle !important;
                font-family: 'Times New Roman', Times, serif !important;
                font-weight: normal !important;
                text-transform: none !important;
                letter-spacing: normal !important;
                min-width: 75px;
                width: 75px;
                max-width: 75px;
                word-break: break-word;
                white-space: normal !important;
                padding-left: 0 !important;
            }
        }
    body, table, th, td, .container, .no-data, button, div, span {
        font-family: 'Times New Roman', Times, serif !important;
        font-size: 16px !important;
    }
    .container {
        max-width: 900px;
        margin: 40px auto;
        background: linear-gradient(135deg, rgba(253,246,238,0.5) 0%, rgba(243,247,250,0.5) 100%);
        padding: 2rem;
        border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.11);
        text-align: center;
        font-family: 'Times New Roman', Times, serif;
        backdrop-filter: blur(8px);
        background: rgba(255,255,255,0.55);
        border: 1px solid rgba(255,255,255,0.18);
    }
    h2 {
        text-align: center;
        font-family: 'Poppins', 'Times New Roman', Times, serif !important;
        font-size: 28px !important;
        font-weight: bold;
    }
    .table-responsive {
        width: 100%;
        margin-top: 2rem;
        /* overflow: visible;  -- ensure dropdowns are not clipped */
    }
    table, .open-defects-table {
        width: auto;
        margin-left: -18px;
        margin-right: auto;
        border-collapse: separate;
        border-spacing: 0;
        border: 1.5px solid #e0c9a6;
        background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%);
        border-radius: 12px;
        box-shadow: 0 4px 16px rgba(180,150,100,0.10), 0 1.5px 4px rgba(180,150,100,0.08);
        overflow: visible;
        table-layout: auto;
    }
    th, td {
        font-size: 13px !important;
        padding: 4px 6px;
        white-space: normal !important;
        word-break: break-word;
        max-width: 180px;
        font-family: 'Times New Roman', Times, serif !important;
        font-weight: normal !important;
        text-transform: none !important;
        overflow: visible !important; /* ensure dropdowns are not clipped */
    }
    th, td {
        border: 1px solid #e0c9a6;
        padding: 8px;
        text-align: center;
        vertical-align: middle;
    }
    th {
        background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important;
        color: #7a5c2e !important;
        font-weight: bold;
        font-size: 17px;
        border-bottom: 3px solid #e0c9a6;
    }
    /* Zebra striping for all tables except header row */
    table tr:not(:first-child):nth-of-type(even), .open-defects-table tr:not(:first-child):nth-of-type(even) {
        background: #f9f3e7 !important;
    }
    table tr:not(:first-child):nth-of-type(odd), .open-defects-table tr:not(:first-child):nth-of-type(odd) {
        background: #fdf6ee !important;
    }
    .no-data { color: #b00; margin: 1rem 0; }
    table tr:not(:first-child):hover, .open-defects-table tr:not(:first-child):hover {
        background-color: #f3e7d1 !important;
        transition: background 0.2s;
    }
    /* Status color classes for Detailed TC Summary */
    .status-blocked { color: #d32f2f !important; font-weight: bold; }
    .status-failed { color: #d32f2f !important; font-weight: bold; }
    .status-in-progress { color: #388e3c !important; font-weight: bold; }
    .status-completed { color: #1976d2 !important; font-weight: bold; }
    .status-execution-completed { color: #1976d2 !important; font-weight: bold; }
    .status-execution-complete-uploaded-for-review { color: #1976d2 !important; font-weight: bold; }
    .status-on-hold { color: #ff9800 !important; font-weight: bold; }
    .status-not-started { color: #757575 !important; font-weight: bold; }
    .status-deferred { color: #757575 !important; font-weight: bold; }
    .status-descoped { color: #757575 !important; font-weight: bold; }
    .status-default { color: #222 !important; font-weight: bold; }
    </style>
</head>
<body style="position:relative; min-height:100vh; background: #f7f3ea;">
<!-- IBM and AT&T Logos -->
<img src="/static/ibm.png" alt="IBM Logo" style="position:fixed; top:24px; left:32px; height:44px; z-index:10; background:transparent; border-radius:8px; padding:4px 10px; box-shadow:0 2px 8px rgba(0,0,0,0.07);">
<img src="/static/att.png" alt="AT&T Logo" style="position:fixed; top:24px; right:32px; height:44px; z-index:10; background:transparent; border-radius:8px; padding:4px 10px; box-shadow:0 2px 8px rgba(0,0,0,0.07);">
<!-- Soft beige wave SVG background -->
<div style="position:fixed; z-index:-1; inset:0; width:100vw; height:100vh; pointer-events:none;">
    <svg width="100%" height="100%" viewBox="0 0 1920 1080" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg">
        <defs>
            <linearGradient id="beigeGrad" x1="0" y1="0" x2="0" y2="1">
                <stop offset="0%" stop-color="#ede3ce"/>
                <stop offset="100%" stop-color="#f7f3ea"/>
            </linearGradient>
        </defs>
        <rect width="1920" height="1080" fill="url(#beigeGrad)"/>
        <path d="M0,200 Q600,350 1920,120 Q1920,0 0,0 Z" fill="#e8dbc0" fill-opacity="0.45"/>
        <path d="M0,600 Q900,900 1920,400 Q1920,0 0,0 Z" fill="#e2d2b6" fill-opacity="0.32"/>
        <path d="M0,900 Q1200,1200 1920,700 Q1920,0 0,0 Z" fill="#e6dcc5" fill-opacity="0.22"/>
    </svg>
</div>
<div class="container" style="position:relative;">
    <a href="/" style="position:absolute; top:22px; right:28px; z-index:2; text-decoration:none;">
        <i class="fa-solid fa-house home-icon-ete"></i>
    </a>
    <style>
    .home-icon-ete {
        font-size: 1.3em;
        color: #7a5c2e;
        background: rgba(255,255,255,0.45);
        border-radius: 12px;
        padding: 8px 12px;
        box-shadow: 0 4px 16px rgba(180,150,100,0.10);
        backdrop-filter: blur(4px);
        transition: background 0.18s, color 0.18s;
        cursor: pointer;
    }
    .home-icon-ete:hover {
        background: rgba(245,233,218,0.85);
        color: #5a3d13;
    }
    </style>
<div style="margin-bottom: 1.2em; position: relative;">
    <h2 style="margin: 0; text-align: center; font-size: 1.5em; font-family: 'Times New Roman', Times, serif !important; font-weight: bold;">{{ dashboard_title|default('ETE Project Dashboard - L1 Report') }}</h2>
</div>
{% if errors %}
    {% for error in errors %}
        <div style="color:red;">{{ error }}</div>
    {% endfor %}
{% endif %}
{% if not rows or rows|length == 0 %}
    <div class="no-data">No data found in any file.</div>
{% else %}
    <div class="table-responsive">
    <!-- Filter box above the table, aligned with Overall Status column -->
    <div style="display: flex; justify-content: flex-end; margin-bottom: 8px;">
        <div style="position: relative; display: inline-block; margin-left: 24px;">
            <input id="globalFilterBox" type="text" placeholder="Filter all columns..." style="padding: 6px 12px; border: 1.2px solid #e0c9a6; border-radius: 6px; font-size: 14px; font-family: 'Times New Roman', Times, serif; width: 120px; background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%); box-shadow: 0 2px 6px rgba(180,150,100,0.07); outline: none; transition: border 0.18s, box-shadow 0.18s;" oninput="document.getElementById('clearFilterBox').style.display = this.value ? 'block' : 'none';resetAllExpandArrows();" />
            <script>
            function resetAllExpandArrows() {
                // Hide all detail rows
                document.querySelectorAll('tr[id^="details-"]').forEach(function(tr) {
                    tr.style.display = 'none';
                });
                // Hide all child detail sections (TC Summary, Detailed TC Summary, Open Defects, etc.)
                document.querySelectorAll('[id^="tc-summary-"], [id^="detailed-tc-summary-"], [id^="open-defects-"]').forEach(function(section) {
                    section.style.display = 'none';
                });
                // Reset all expand arrows to collapsed (right)
                document.querySelectorAll('[id^="arrow-"]').forEach(function(arrowEl) {
                    var svg = arrowEl.querySelector('svg');
                    if (svg) svg.style.transform = 'rotate(0deg)';
                });
            }
            </script>
            <span id="clearFilterBox" style="display:none; position: absolute; right: 8px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #bfa16a; font-size: 16px; user-select: none;" onclick="var f=document.getElementById('globalFilterBox');f.value='';f.dispatchEvent(new Event('input'));this.style.display='none';">&#10005;</span>
        </div>
    </div>
<table id="mainTable">
            <tr>
                {% for col in columns %}
                <th style="position: relative; white-space:nowrap; padding: 8px 10px;{% if col|string|trim|lower == 'overall status' %} min-width:130px;{% elif col|string|trim|lower == 'plan end' %} min-width:60px;{% endif %}" class="{% if col|string|trim|lower == 'overall status' %}status-col{% endif %}">
                    <div style="display:flex; justify-content:center; align-items:center; gap:2px; margin-top:2px;">
                        <span class="sort-icons" style="cursor:pointer;display:inline-flex;flex-direction:column;align-items:center;gap:1px; margin-right:2px;" onclick="sortTable({{ loop.index0 }})">
                            <svg width="12" height="8" viewBox="0 0 16 10" style="display:block;">
                                <polygon points="8,2 3,8 13,8" fill="#222"/>
                            </svg>
                            <svg width="12" height="8" viewBox="0 0 16 10" style="display:block;">
                                <polygon points="8,8 3,2 13,2" fill="#222"/>
                            </svg>
                        </span>
                        <div style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; font-weight:600; font-size:{% if loop.index0 == 0 %}12px{% else %}1em{% endif %};">{{ col }}</div>
                        <!-- Filter icon removed -->
                    </div>
                    <div class="column-filter-dropdown" data-col="{{ loop.index0 }}" style="display:none; position: absolute; top: 100%; left: 0; background: none; border: none; border-radius: 9px; box-shadow: none; padding: 6px 10px; z-index: 10; min-width: 110px;">
                        <input type="text" class="column-filter-input modern-filter-input" placeholder="Type to filter..." style="width: 100%; padding: 5px 9px; border: 1.2px solid #e0c9a6; border-radius: 6px; font-size: 13px; font-family: 'Times New Roman', Times, serif; background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%); box-shadow: 0 2px 6px rgba(180,150,100,0.07); outline: none; transition: border 0.18s, box-shadow 0.18s;">
                    </div>
                </th>
                {% endfor %}
            </tr>
            {% for row in rows %}
            {% set row_idx = loop.index0 %}
            <tr data-rowid="details-{{ row_idx }}">
                {% for cell in row.cells %}
                    {% set cell_idx = loop.index0 %}
                    {% set display_cell = cell if cell|string|trim != '' else 'N/A' %}
                    {% if cell_idx == 0 %}
                    <td style="vertical-align: middle; text-align: center; height: 100%; font-size:14px !important;">
                        <span style="display: flex; align-items: center; justify-content: flex-start; width: 100%; height: 100%; font-size: 14px !important; font-family: 'Times New Roman', Times, serif; font-weight: normal; text-align: left;">
                            <button onclick="toggleDetails('details-{{ row_idx }}')" style="background:none; border:none; cursor:pointer; font-size:1.2em; padding:0 4px 0 0; margin-right:4px; display: flex; align-items: center;">
                                <span id="arrow-{{ row_idx }}" style="display:inline-block; color:#1976d2; vertical-align:middle;">
                                    <svg width="20" height="20" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;">
                                        <polyline points="7,5 13,10 7,15" stroke="#1976d2" stroke-width="2.5" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
                                    </svg>
                                </span>
                            </button>
                            <span style="display: flex; align-items: center; font-size:14px !important;">{{ display_cell }}</span>
                        </span>
                    </td>
                    {% elif status_col_idx is not none and cell_idx == status_col_idx and cell is mapping %}
                    {% set status_val = cell.value|string|trim|upper %}
                    <td class="status-col" style="background: transparent; color: {{ cell.font_color if cell.font_color else 'black' }}; text-align:center; vertical-align:middle;">
                        <span style="
                            display: inline-block;
                            border-radius: 16px;
                            padding: 6px 18px;
                            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
                            border: 1.5px solid #fff;
                            min-width: 90px;
                            background-color: {{ cell.bg_color if cell.bg_color and cell.bg_color|lower not in ['#1976d2', '#b7d7f7', '#e3f0fc', '#a4c8ea', '#007bff', '#154360'] else '' }};
                            color: #fff !important;
                            font-family: 'Times New Roman', Times, serif !important;
                            font-size: 13px !important;
                            font-weight: bold !important;
                            text-transform: none !important;
                            letter-spacing: normal !important;
                        ">
                            {{ cell.value if cell.value|string|trim != '' else 'N/A' }}
                        </span>
                    </td>
                    {% else %}
                    <td>{{ display_cell }}</td>
                    {% endif %}
                {% endfor %}
            </tr>
            <tr id="details-{{ row_idx }}" style="display:none;">
                <td colspan="{{ columns|length + 1 }}" style="padding-top: 0; text-align: center; background:#fdf6ee !important; border-radius:0 !important; box-shadow:none !important;">
                    <div style="height: 1px; background: #e0e0e0; margin: 0 0 18px 0;"></div>
                    <!-- TC Summary Expandable -->
                    <div style="margin-bottom: 1em; text-align: center;">
                        <button onclick="toggleChildDetails('tc-summary-{{ row_idx }}')" style="background:none; border:none; cursor:pointer; font-size:1.1em; margin-right:6px; color:#1976d2;">
                            <span id="arrow-tc-summary-{{ row_idx }}" style="display:inline-block; color:#1976d2; vertical-align:middle;">
                                <svg width="16" height="16" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;">
                                    <polyline points="7,5 13,10 7,15" stroke="#1976d2" stroke-width="2.5" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
                                </svg>
                            </span>
                        </button>
                        <b><u style="font-size:1.35em;">TC Summary</u></b>
                    </div>
                    <div id="tc-summary-{{ row_idx }}" style="display:none; padding:1.2em 1.5em; background:#fdf6ee !important; border:none !important; box-shadow:none !important; font-size:14px; text-align:center; vertical-align:middle; border-radius: 12px; margin-bottom: 1em;">
                        {% if row.tc_summary and row.tc_summary|length > 0 %}
                            {% set tc_map = {} %}
                            {% set tc_labels = [
                                'Completed',
                                'Execution Complete/Uploaded for review',
                                'Blocked',
                                'Failed',
                                'On Hold',
                                'In Progress',
                                'Not Started',
                                'Deferred',
                                'Descoped',
                                'Total TCs'] %}
                            {% for row_tc in row.tc_summary %}
                                {% if row_tc|length > 1 and row_tc[0] in tc_labels %}
                                    {% set _ = tc_map.update({row_tc[0]: row_tc[1]}) %}
                                {% endif %}
                            {% endfor %}
                            <div style="margin-top:0.5em;">
                                <table style="margin:auto; border-collapse:collapse; font-size:14px; width: 480px; min-width: 480px;">
                                                                    <style>
                                                                        #tc-summary-{{ row_idx }} td:first-child {
                                                                            text-align: center !important;
                                                                        }
                                                                    </style>
                                <style>
                                    #tc-summary-{{ row_idx }} table {
                                        border-radius: 16px !important;
                                        overflow: hidden;
                                    }
                                </style>
                                    <tr>
                                        <th style="padding:6px 24px; text-align:center; width:260px; font-weight:bold !important; font-size:16px !important; white-space:nowrap;">Status</th>
                                        <th style="padding:6px 24px; text-align:center; width:120px; font-weight:bold !important; font-size:16px !important;">Count</th>
                                    </tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Completed</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#0a2348; font-weight:normal !important;">{{ tc_map['Completed']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Execution Complete/Uploaded for review</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#0a2348; font-weight:normal !important;">{{ tc_map['Execution Complete/Uploaded for review']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Blocked</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#d32f2f; font-weight:normal !important;">{{ tc_map['Blocked']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Failed</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#d32f2f; font-weight:normal !important;">{{ tc_map['Failed']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">On Hold</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#ff9800; font-weight:normal !important;">{{ tc_map['On Hold']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">In Progress</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#388e3c; font-weight:normal !important;">{{ tc_map['In Progress']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Not Started</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#757575; font-weight:normal !important;">{{ tc_map['Not Started']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Deferred</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#757575; font-weight:normal !important;">{{ tc_map['Deferred']|default('0') }}</span></td></tr>
                                    <tr><td style="font-weight:normal !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap;">Descoped</td><td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important;"><span style="color:#757575; font-weight:normal !important;">{{ tc_map['Descoped']|default('0') }}</span></td></tr>
                                    <tr>
                                        <td style="font-weight:bold !important; width:260px; padding:6px 24px; text-align:center; vertical-align:middle; font-size:14px !important; white-space:nowrap; background:#ffe082;">Total TCs</td>
                                        <td style="text-align:center; width:120px; padding:6px 24px; font-size:14px !important; font-weight:bold !important; background:#ffe082;">
                                            <span style="color:#222; font-weight:bold !important;">{{ tc_map['Total TCs']|default('0') }}</span>
                                        </td>
                                    </tr>
                                </table>
                            </div>
                        {% else %}
                        <div style="color:#b00;">No TC Summary found.</div>
                        {% endif %}
                    </div>
                    <!-- Detailed TC Summary Expandable -->
                    <div style="margin-bottom: 1em; margin-top: 1em;">
                        <div style="text-align:center;">
                            <button onclick="toggleChildDetails('detailed-tc-summary-{{ row_idx }}')" style="background:none; border:none; cursor:pointer; font-size:1.1em; margin-right:6px; color:#1976d2;">
                                <span id="arrow-detailed-tc-summary-{{ row_idx }}" style="display:inline-block; color:#1976d2; vertical-align:middle;">
                                    <svg width="16" height="16" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;">
                                        <polyline points="7,5 13,10 7,15" stroke="#1976d2" stroke-width="2.5" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
                                    </svg>
                                </span>
                            </button>
                            <b><u style="font-size:1.35em;">Detailed TC Summary</u></b>
                        </div>
                    </div>
                    {% macro status_class(status) %}
                        {%- set s = status|string|trim|upper -%}
                        {%- if s == 'BLOCKED' -%}status-blocked
                        {%- elif s == 'FAILED' -%}status-failed
                        {%- elif s == 'IN PROGRESS' -%}status-in-progress
                        {%- elif s == 'COMPLETED' -%}status-completed
                        {%- elif s == 'EXECUTION COMPLETED' -%}status-execution-completed
                        {%- elif s == 'EXECUTION COMPLETE/UPLOADED FOR REVIEW' -%}status-execution-complete-uploaded-for-review
                        {%- elif s == 'ON HOLD' -%}status-on-hold
                        {%- elif s == 'NOT STARTED' -%}status-not-started
                        {%- elif s == 'DEFERRED' -%}status-deferred
                        {%- elif s == 'DESCOPED' -%}status-descoped
                        {%- else -%}status-default
                        {%- endif -%}
                    {% endmacro %}
                    {% macro status_color(status) %}
                        {%- set s = status|string|trim|upper -%}
                        {%- if s == 'BLOCKED' or s == 'FAILED' -%}#d32f2f
                        {%- elif s == 'IN PROGRESS' -%}#388e3c
                        {%- elif s == 'COMPLETED' or s == 'EXECUTION COMPLETED' or s == 'EXECUTION COMPLETE/UPLOADED FOR REVIEW' -%}#1976d2
                        {%- elif s == 'ON HOLD' -%}#ff9800
                        {%- elif s == 'NOT STARTED' or s == 'DEFERRED' or s == 'DESCOPED' -%}#757575
                        {%- else -%}#222
                        {%- endif -%}
                    {% endmacro %}
                    <div id="detailed-tc-summary-{{ row_idx }}" style="display:none !important; padding:1.2em 1.5em; background:#fdf6ee !important; border:none !important; box-shadow:none !important; font-size:14px; text-align:center; vertical-align:middle; border-radius: 12px;">
                        {% if row.detailed_tc_headers and row.detailed_tc_data and row.detailed_tc_data|length > 0 %}
                            <table class="open-defects-table" style="margin:auto;">
                                <style>
                                    #detailed-tc-summary-{{ row_idx }} table, #detailed-tc-summary-{{ row_idx }} th, #detailed-tc-summary-{{ row_idx }} td {
                                        font-size: 10px !important;
                                    }
                                </style>
                                <tr>
                                    {% for header in row.detailed_tc_headers %}
                                        {# TC No. #}
                                        {% if loop.index0 == 0 %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:50px; width:50px;">{{ header }}</th>
                                        {# TC Name #}
                                        {% elif loop.index0 == 1 %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:200px; width:240px;">{{ header }}</th>
                                        {# Status #}
                                        {% elif loop.index0 == 2 %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:65px; width:65px;">{{ header }}</th>
                                        {# Start Date, End Date #}
                                        {% elif loop.index0 == 3 or loop.index0 == 4 %}
                                                <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:60px; width:60px;">{{ header }}</th>
                                        {# Planned Steps, Actual Steps Passed, Planned %, Actual % #}
                                            {% elif header|string|trim|lower in ['total steps', 'planned steps', 'actual steps passed', 'planned %', 'actual %'] %}
                                                <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:50px; width:50px;">{{ header }}</th>
                                        {# Default #}
                                        {% else %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center;">{{ header }}</th>
                                        {% endif %}
                                    {% endfor %}
                                </tr>
                                {% set status_idx = None %}
                                {% for h in row.detailed_tc_headers %}
                                    {% if h|string|trim|lower == 'status' %}{% set status_idx = loop.index0 %}{% endif %}
                                {% endfor %}
                                {% for tc_row in row.detailed_tc_data %}
                                    <tr>
                                        {# Use a namespace to persist total_idx #}
                                        {% set ns = namespace(total_idx=None) %}
                                        {% for c in tc_row %}
                                            {% if c|string|trim|lower == 'total' %}
                                                {% set ns.total_idx = loop.index0 %}
                                            {% endif %}
                                        {% endfor %}
                                        {% for cell in tc_row %}
                                            {% if ns.total_idx is not none and loop.index0 >= ns.total_idx %}
                                                <td style="text-align:center !important; font-weight:bold !important; background:#ffe082; min-width:50px; font-size:12px !important;">{{ cell }}</td>
                                            {% elif loop.index0 == 0 %}
                                                <td style="text-align:center !important; min-width:50px; width:50px; font-weight:bold !important;">{{ cell }}</td>
                                            {% elif loop.index0 == 1 %}
                                                <td style="text-align:center !important; min-width:200px; width:240px;">{{ cell }}</td>
                                            {% elif loop.index0 == 2 %}
                                                <td style="text-align:center !important; min-width:70px; width:70px;"><span style="color: {{ status_color(cell) }}; font-weight:bold; font-size:9px !important;">{{ cell }}</span></td>
                                            {% elif loop.index0 == 3 or loop.index0 == 4 %}
                                                <td style="text-align:center !important; min-width:60px; width:60px; font-size:12px !important;">{{ cell.split(' ')[0] if cell and ' ' in cell else cell }}</td>
                                            {% elif header and header[loop.index0]|string|trim|lower in ['total steps', 'planned steps', 'actual steps passed', 'planned %', 'actual %'] %}
                                                <td style="text-align:center !important; min-width:50px; width:50px; font-size:12px !important;">{{ cell }}</td>
                                            {% else %}
                                                <td style="text-align:center !important;">{{ cell }}</td>
                                            {% endif %}
                                        {% endfor %}
                                    </tr>
                                {% endfor %}
                            </table>
                        {% else %}
                        <div style="color:#b00;">No Detailed TC Summary found.</div>
                        {% endif %}
                    </div>
                    <!-- Open Defects Expandable -->
                    <div style="margin-bottom: 1em; margin-top: 1em;">
                        <div style="text-align:center;">
                            <button onclick="toggleChildDetails('open-defects-{{ row_idx }}')" style="background:none; border:none; cursor:pointer; font-size:1.1em; margin-right:6px; color:#1976d2;">
                                <span id="arrow-open-defects-{{ row_idx }}" style="display:inline-block; color:#1976d2; vertical-align:middle;">
                                    <svg width="16" height="16" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;">
                                        <polyline points="7,5 13,10 7,15" stroke="#1976d2" stroke-width="2.5" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
                                    </svg>
                                </span>
                            </button>
                            <b><u style="font-size:1.35em;">Open Defects</u></b>
                        </div>
                    </div>
                    <div id="open-defects-{{ row_idx }}" style="display:none; padding:1.2em 1.5em; background:#f6f8fa; border:1px solid #e0e0e0; font-size:14px; text-align:center; vertical-align:middle; border-radius: 12px; box-shadow: 0 2px 12px rgba(44,62,80,0.10);">
                        {% if row.open_defects and row.open_defects|length > 0 %}
                            <table class="open-defects-table" style="margin:auto; font-size:12px !important;">
                                <style>
                                    #open-defects-{{ row_idx }} table, #open-defects-{{ row_idx }} th, #open-defects-{{ row_idx }} td {
                                        font-size: 12px !important;
                                    }
                                </style>
                                {% set od_header = row.open_defects[0] %}
                                {% set status_idx = None %}
                                {% for col in od_header %}
                                    {% if col|string|lower == 'status' %}
                                        {% set status_idx = loop.index0 %}
                                    {% endif %}
                                {% endfor %}
                                <tr>
                                    {% for val in od_header %}
                                        {% if val|string|lower == 'defect id' %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:90px; width:90px; max-width:90px; background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important; color: #7a5c2e !important; border-bottom: 3px solid #e0c9a6;">{{ val if val|string|trim != '' else 'N/A' }}</th>
                                        {% elif val|string|lower == 'comments' %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:340px; width:340px; max-width:540px; background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important; color: #7a5c2e !important; border-bottom: 3px solid #e0c9a6;">{{ val if val|string|trim != '' else 'N/A' }}</th>
                                        {% elif val|string|lower == 'severity' %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; min-width:42px; width:42px; max-width:42px; background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important; color: #7a5c2e !important; border-bottom: 3px solid #e0c9a6;">{{ val if val|string|trim != '' else 'N/A' }}</th>
                                        {% else %}
                                            <th style="font-weight:bold !important; font-size:12px !important; text-align:center; background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important; color: #7a5c2e !important; border-bottom: 3px solid #e0c9a6;">{{ val if val|string|trim != '' else 'N/A' }}</th>
                                        {% endif %}
                                    {% endfor %}
                                </tr>
                                {% for defect_row in row.open_defects[1:] %}
                                    {% if status_idx is not none and defect_row|length > status_idx %}
                                        {% set defect_status = defect_row[status_idx]|string|trim|lower %}
                                        {% if defect_status != 'accepted' %}
                                            <tr>
                                            {% for val in defect_row %}
                                                {% set idx = loop.index0 %}
                                                {% if od_header[idx]|string|lower == 'defect id' %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc; text-align:center !important; vertical-align:middle !important; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; font-size:1em; font-weight:bold; min-width:90px; width:90px; max-width:90px;">
                                                        {{ (val if val|string|trim != '' else 'N/A')|replace('___LINEBREAK___', '<br>')|safe }}
                                                    </td>
                                                {% elif od_header[idx]|string|lower == 'severity' %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc; min-width:42px; width:42px; max-width:42px; text-align:center;">{{ (val if val|string|trim != '' else 'N/A')|replace('___LINEBREAK___', '<br>')|safe }}</td>
                                                {% elif od_header[idx]|string|lower == 'comments' %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc; min-width:340px; width:340px; max-width:540px;">{{ (val if val|string|trim != '' else 'N/A')|replace('___LINEBREAK___', '<br>')|safe }}</td>
                                                {% else %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc;">{{ (val if val|string|trim != '' else 'N/A')|replace('___LINEBREAK___', '<br>')|safe }}</td>
                                                {% endif %}
                                            {% endfor %}
                                            </tr>
                                        {% endif %}
                                    {% elif status_idx is none %}
                                        <tr>
                                        {% for val in defect_row %}
                                            {% set idx = loop.index0 %}
                                            {% if od_header[idx]|string|lower == 'defect id' %}
                                                <td style="padding: 0.3em 1em; border: 1px solid #ccc; text-align:center !important; vertical-align:middle !important; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; font-size:1em; font-weight:bold; min-width:90px; width:90px; max-width:90px;">
                                                    {{ (val if val|string|trim != '' else 'N/A')|replace('___LINEBREAK___', '<br>')|safe }}
                                                </td>
                                                {% elif od_header[idx]|string|lower == 'defect opened' %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc;">
                                                        {%- set cell_val = val if val|string|trim != '' else 'N/A' -%}
                                                        {{ cell_val.split(' ')[0] if ' ' in cell_val else cell_val }}
                                                    </td>
                                                {% elif od_header[idx]|string|lower == 'comments' %}
                                                    <td style="padding: 0.3em 1em; border: 1px solid #ccc; min-width:340px; width:340px; max-width:540px;">
                                                        {%- set cell_val = val if val|string|trim != '' else 'N/A' -%}
                                                        {%- if idx == od_header|length - 1 and 'comments' in od_header[idx]|string|lower -%}
                                                            {{ cell_val|replace('___LINEBREAK___', '<br>')
                                                                |replace('Description', '<b>Description</b><br>')
                                                                |replace('Blocker', '<br><b>Blocker</b><br>')
                                                                |replace('Comments', '<br><b>Comments</b><br>')|safe }}
                                                        {%- else -%}
                                                            {{ cell_val|replace('___LINEBREAK___', '<br>')|safe }}
                                                        {%- endif -%}
                                                    </td>
                                            {% else %}
                                                <td style="padding: 0.3em 1em; border: 1px solid #ccc;">
                                                    {%- set cell_val = val if val|string|trim != '' else 'N/A' -%}
                                                    {%- if idx == od_header|length - 1 and 'comments' in od_header[idx]|string|lower -%}
                                                        {{ cell_val|replace('___LINEBREAK___', '<br>')
                                                            |replace('Description', '<b>Description</b>')
                                                            |replace('Blocker', '<b>Blocker</b>')
                                                            |replace('Comments', '<b>Comments</b>')|safe }}
                                                    {%- else -%}
                                                        {{ cell_val|replace('___LINEBREAK___', '<br>')|safe }}
                                                    {%- endif -%}
                                                </td>
                                            {% endif %}
                                        {% endfor %}
                                        </tr>
                                    {% endif %}
                                {% endfor %}
                            </table>
                        {% else %}
                        <div style="color:#b00;">No Open Defects found.</div>
                        {% endif %}
                    </div>
                </td>
            </tr>
            {% endfor %}
    </table>
    </div>
    <script>
                // --- Sort Table Function ---
                function sortTable(colIdx) {
                    var table = document.getElementById('mainTable');
                    var tbody = table.querySelector('tbody');
                    if (!tbody) return;
                    // Only sort direct children of tbody that are not detail rows
                    var rows = Array.from(tbody.children).filter(function(tr) {
                        return tr.tagName === 'TR' && tr.querySelectorAll('td').length && !tr.id.startsWith('details-');
                    });
                    // Track sort direction
                    if (!table._sortState) table._sortState = {};
                    var state = table._sortState;
                    state[colIdx] = state[colIdx] === 'asc' ? 'desc' : 'asc';
                    var dir = state[colIdx];
                    rows.sort(function(a, b) {
                        var cellA = a.querySelectorAll('td')[colIdx];
                        var cellB = b.querySelectorAll('td')[colIdx];
                        var valA = cellA ? cellA.textContent.trim() : '';
                        var valB = cellB ? cellB.textContent.trim() : '';
                        // Try to compare as numbers, fallback to string
                        var numA = parseFloat(valA.replace(/[^0-9.\-]/g, ''));
                        var numB = parseFloat(valB.replace(/[^0-9.\-]/g, ''));
                        if (!isNaN(numA) && !isNaN(numB)) {
                            return dir === 'asc' ? numA - numB : numB - numA;
                        } else {
                            valA = valA.toLowerCase();
                            valB = valB.toLowerCase();
                            if (valA < valB) return dir === 'asc' ? -1 : 1;
                            if (valA > valB) return dir === 'asc' ? 1 : -1;
                            return 0;
                        }
                    });
                    // Hide all detail rows after sorting
                    Array.from(tbody.querySelectorAll('tr[id^="details-"]')).forEach(function(tr) {
                        tr.style.display = 'none';
                    });
                    // Remove all main rows (not detail rows) from tbody
                    rows.forEach(function(tr) {
                        tbody.removeChild(tr);
                    });
                    // Re-append sorted main rows at the top of tbody (before any detail rows)
                    var firstDetailRow = Array.from(tbody.children).find(function(tr) {
                        return tr.id && tr.id.startsWith('details-');
                    });
                    rows.forEach(function(tr) {
                        if (firstDetailRow) {
                            tbody.insertBefore(tr, firstDetailRow);
                        } else {
                            tbody.appendChild(tr);
                        }
                    });
                }
        // --- Filter Icon Highlighting ---
        function updateFilterIconHighlight() {
            // Main table
            document.querySelectorAll('.column-filter-dropdown').forEach(function(dropdown) {
                var colIdx = dropdown.getAttribute('data-col');
                var input = dropdown.querySelector('input');
                var icon = dropdown.parentElement.querySelector('.filter-svg');
                if (input && icon) {
                    if (input.value.trim() !== '') {
                        icon.querySelector('path').setAttribute('stroke', '#333');
                    } else {
                        icon.querySelector('path').setAttribute('stroke', '#888');
                    }
                }
            });
            // Open Defects tables
            document.querySelectorAll('.od-column-filter-dropdown').forEach(function(dropdown) {
                var input = dropdown.querySelector('input');
                var icon = dropdown.parentElement.querySelector('.od-filter-svg');
                if (input && icon) {
                    if (input.value.trim() !== '') {
                        icon.querySelector('path').setAttribute('stroke', '#333');
                    } else {
                        icon.querySelector('path').setAttribute('stroke', '#888');
                    }
                }
            });
        }
    // Column Filter Dropdown Logic for Main Table
    function toggleColumnFilter(icon, colIdx, event) {
        // Hide all other filter dropdowns
        document.querySelectorAll('.column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
        // Find the dropdown for this column (relative to the <th>)
        var th = icon.closest('th');
        var dropdown = th ? th.querySelector('.column-filter-dropdown') : null;
        if (dropdown) {
            // Show the dropdown
            dropdown.style.display = 'block';
            var input = dropdown.querySelector('input');
            if (input) { input.focus(); }
        }
        // Prevent click from closing the dropdown
        if (event) event.stopPropagation();
    }
    document.addEventListener('click', function() {
        document.querySelectorAll('.column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
        document.querySelectorAll('.od-column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
    });
    document.querySelectorAll('.column-filter-input').forEach(function(input) {
        input.addEventListener('input', function() {
            var colIdx = parseInt(input.parentElement.getAttribute('data-col'));
            var filter = input.value.toLowerCase();
            var table = document.getElementById('mainTable');
            var trs = table.querySelectorAll('tbody > tr');
            trs.forEach(function(tr) {
                if (tr.querySelectorAll('td').length && !tr.id.startsWith('details-')) {
                    var tds = tr.querySelectorAll('td');
                    var cell = tds[colIdx];
                    if (cell) {
                        var text = cell.textContent.toLowerCase();
                        tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                    }
                }
            });
            updateFilterIconHighlight();
        });
    });

    // Column Filter Dropdown Logic for Open Defects Table
    function toggleODColumnFilter(icon, colIdx, rowIdx) {
        document.querySelectorAll('.od-column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
        var dropdown = icon.parentElement.querySelector('.od-column-filter-dropdown');
        if (dropdown) {
            dropdown.style.display = dropdown.style.display === 'none' ? 'block' : 'none';
            var input = dropdown.querySelector('input');
            if (input) { input.focus(); }
        }
        event.stopPropagation();
    }
    document.querySelectorAll('.od-column-filter-input').forEach(function(input) {
        input.addEventListener('input', function() {
            var colIdx = parseInt(input.parentElement.getAttribute('data-col'));
            var rowIdx = parseInt(input.parentElement.getAttribute('data-row'));
            var table = document.querySelector(`#open-defects-${rowIdx} table`);
            if (!table) return;
            var filter = input.value.toLowerCase();
            var trs = table.querySelectorAll('tr');
            trs.forEach(function(tr, idx) {
                if (idx === 0) return; // header
                var tds = tr.querySelectorAll('td');
                var cell = tds[colIdx];
                if (cell) {
                    var text = cell.textContent.toLowerCase();
                    tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                }
            });
            updateFilterIconHighlight();
        });
    });
    // Main Table Filter
    document.addEventListener('DOMContentLoaded', function() {
                updateFilterIconHighlight();
                // Also update on page load in case filters are pre-filled
                document.querySelectorAll('.column-filter-input').forEach(function(input) {
                    input.addEventListener('change', updateFilterIconHighlight);
                });
                document.querySelectorAll('.od-column-filter-input').forEach(function(input) {
                    input.addEventListener('change', updateFilterIconHighlight);
                });
        var mainFilter = document.getElementById('mainTableFilter');
        if (mainFilter) {
            mainFilter.addEventListener('input', function() {
                var filter = mainFilter.value.toLowerCase();
                var table = document.getElementById('mainTable');
                var trs = table.querySelectorAll('tbody > tr');
                trs.forEach(function(tr) {
                    // Only filter main rows, not details rows
                    if (tr.querySelectorAll('td').length && !tr.id.startsWith('details-')) {
                        var text = tr.textContent.toLowerCase();
                        tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                    }
                });
            });
        }
        // Open Defects Table Filters
        document.querySelectorAll('.openDefectsFilter').forEach(function(input) {
            input.addEventListener('input', function() {
                var filter = input.value.toLowerCase();
                var table = input.parentElement.nextElementSibling;
                if (!table) return;
                var trs = table.querySelectorAll('tr');
                trs.forEach(function(tr, idx) {
                    // Always show header row
                    if (idx === 0) return;
                    var text = tr.textContent.toLowerCase();
                    tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                });
            });
        });
    });
    function toggleDetails(id) {
        var row = document.getElementById(id);
        var arrow = document.getElementById('arrow-' + id.split('-')[1]);
        var parentRow = document.getElementById('mainTable').querySelector('tr[data-rowid="' + id + '"]');
        var isOpen = row && row.style.display !== 'none';
        // If already open, collapse it and reset arrow
        if (isOpen) {
            row.style.display = 'none';
            if (arrow && arrow.querySelector('svg')) {
                arrow.querySelector('svg').style.transform = 'rotate(0deg)';
            }
            return;
        }
        // Otherwise, collapse all others and open this one
        document.querySelectorAll('tr[id^="details-"]').forEach(function(tr) {
            tr.style.display = 'none';
        });
        document.querySelectorAll('[id^="arrow-"]').forEach(function(arrowEl) {
            if (arrowEl.querySelector('svg')) {
                arrowEl.querySelector('svg').style.transform = 'rotate(0deg)';
            }
        });
        // Move the detail row right after its parent row if not already there
        if (row && parentRow && row.previousElementSibling !== parentRow) {
            parentRow.parentNode.insertBefore(row, parentRow.nextElementSibling);
        }
        if (row) {
            row.style.display = '';
            if (arrow && arrow.querySelector('svg')) {
                arrow.querySelector('svg').style.transform = 'rotate(90deg)';
            }
        }
    }
    function toggleChildDetails(id) {
        var section = document.getElementById(id);
        var arrow = document.getElementById('arrow-' + id);
        if (section.style.display === 'none') {
            section.style.display = '';
            if (arrow) {
                arrow.querySelector('svg').style.transform = 'rotate(90deg)';
            }
        } else {
            section.style.display = 'none';
            if (arrow) {
                arrow.querySelector('svg').style.transform = 'rotate(0deg)';
            }
        }
    }
    // Global filter for all columns
    var globalFilterBox = document.getElementById('globalFilterBox');
    if (globalFilterBox) {
        globalFilterBox.addEventListener('input', function() {
            var filter = globalFilterBox.value.toLowerCase();
            var table = document.getElementById('mainTable');
            if (!table) return;
            var trs = table.querySelectorAll('tr[data-rowid]');
            trs.forEach(function(tr) {
                var text = tr.textContent.toLowerCase();
                var rowId = tr.getAttribute('data-rowid');
                var detailsRow = document.getElementById(rowId);
                if (text.indexOf(filter) > -1) {
                    tr.style.display = '';
                    if (detailsRow) detailsRow.style.display = 'none'; // Hide details by default on filter
                } else {
                    tr.style.display = 'none';
                    if (detailsRow) detailsRow.style.display = 'none';
                }
            });
        });
    }
    // ...existing code...
    </script>
{% endif %}
</div>
</body>
</html>
'''

# --- CONFIG ---
EXCEL_SHEET = 'Overall Status'  # Used for import only


def _ensure_dashboard_table(table_name: str) -> None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {table_name} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT UNIQUE,
                project_name TEXT,
                columns TEXT,
                cells TEXT,
                tc_summary TEXT,
                open_defects TEXT,
                detailed_tc_headers TEXT,
                detailed_tc_data TEXT
            )'''
        )


def _parse_dashboard_filename(filename: str) -> tuple[str, str, str, int, int] | None:
    """Parse PID dashboard filename.

    Expected patterns:
      PID_<pid>_<project_name>_ETE_Status_as_of_<MM>_<DD>.xlsx
      PID_<pid>_<project_name>_PVT_Status_as_of_<MM>_<DD>.xlsx
      PID_<pid>_<project_name>_ETE_Billing_Status_as_of_<MM>_<DD>.xlsx
    """
    pattern = re.compile(
        r'^PID_(\d+)_(.+?)_(ETE_Billing|ETE|PVT)_Status_as_of_(\d{2})_(\d{2})',
        re.IGNORECASE,
    )
    m = pattern.search(filename or '')
    if not m:
        return None
    pid = m.group(1)
    project_name_raw = m.group(2)
    tag = m.group(3)
    mm = int(m.group(4))
    dd = int(m.group(5))
    project_name = project_name_raw.replace('_', ' ').strip()
    # Normalize tag casing
    tag_norm = tag.strip()
    if tag_norm.lower() == 'ete_billing':
        tag_norm = 'ETE_Billing'
    else:
        tag_norm = tag_norm.upper()
    return pid, project_name, tag_norm, mm, dd


def _table_for_tag(tag: str) -> str:
    tag_norm = (tag or '').strip()
    if tag_norm == 'PVT':
        return 'dashboard_data_pvt'
    if tag_norm in ('ETE_Billing', 'ETE BILLING'):
        return 'dashboard_data_billing'
    return 'dashboard_data'


def _deduplicate_table_latest_by_pid_project(table_name: str) -> int:
    """Keep only the latest (MM,DD) record per (PID, project_name) in the given table.

    Latest-ness is determined from the filename pattern `_Status_as_of_MM_DD`.
    Returns number of deleted rows.
    """
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id, filename, project_name FROM {table_name}")
        rows = cur.fetchall()

        latest: dict[tuple[str, str], tuple[int, int, int]] = {}
        # key -> (id, mm, dd)
        for row_id, filename, project_name in rows:
            parsed = _parse_dashboard_filename(filename or '')
            if not parsed:
                continue
            pid, pname, _tag, mm, dd = parsed
            key = (str(pid).strip(), str(pname or project_name or '').strip())

            if key not in latest or (mm, dd) > (latest[key][1], latest[key][2]):
                latest[key] = (int(row_id), int(mm), int(dd))

        keep_ids = {v[0] for v in latest.values()}
        if not keep_ids:
            return 0

        placeholders = ','.join(['?'] * len(keep_ids))
        cur.execute(
            f"DELETE FROM {table_name} WHERE id NOT IN ({placeholders})",
            tuple(keep_ids),
        )
        deleted = cur.rowcount if cur.rowcount is not None else 0
        conn.commit()
        return deleted


def _import_from_github_routed(*, env_prefix: str, tags_filter: set[str] | None = None) -> tuple[str, int]:
    import requests
    from openpyxl import load_workbook

    token = os.environ.get(f'{env_prefix}_GITHUB_TOKEN') or os.environ.get('GITHUB_TOKEN')
    user = os.environ.get(f'{env_prefix}_GITHUB_USER') or os.environ.get('GITHUB_USER')
    repo = os.environ.get(f'{env_prefix}_GITHUB_REPO') or os.environ.get('GITHUB_REPO')

    if not (token and user and repo):
        return (
            f'Missing GitHub credentials. Set {env_prefix}_GITHUB_TOKEN/{env_prefix}_GITHUB_USER/{env_prefix}_GITHUB_REPO (or the generic GITHUB_TOKEN/GITHUB_USER/GITHUB_REPO).',
            403,
        )

    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://api.github.com/repos/{user}/{repo}/contents"

    r = requests.get(url, headers=headers)
    r.raise_for_status()
    files = r.json()
    excel_files = [f for f in files if f.get('name', '').lower().endswith(('.xlsx', '.xls'))]

    latest_file_map: dict[tuple[str, str, str], tuple[tuple[int, int], dict]] = {}
    for f in excel_files:
        filename = f.get('name', '')
        parsed = _parse_dashboard_filename(filename)
        if not parsed:
            continue
        pid, project_name, tag, mm, dd = parsed
        if tags_filter and tag not in tags_filter:
            continue
        key = (pid, project_name, tag)
        if key not in latest_file_map or (mm, dd) > latest_file_map[key][0]:
            latest_file_map[key] = ((mm, dd), f)

    latest_files = [v[1] for v in latest_file_map.values()]

    imported_counts: dict[str, int] = {
        'dashboard_data': 0,
        'dashboard_data_pvt': 0,
        'dashboard_data_billing': 0,
    }

    for file_info in latest_files:
        download_url = file_info.get('download_url')
        filename = file_info.get('name', '')
        if not download_url or not filename:
            continue

        r = requests.get(download_url, headers=headers)
        if r.status_code != 200:
            continue

        try:
            parsed = _parse_dashboard_filename(filename)
            if not parsed:
                continue
            pid, project_name, tag, mm, dd = parsed
            if tags_filter and tag not in tags_filter:
                continue

            table_name = _table_for_tag(tag)
            _ensure_dashboard_table(table_name)

            df = pd.read_excel(io.BytesIO(r.content), sheet_name=EXCEL_SHEET, header=1)
            selected_cols = [col for col in df.columns if not str(col).startswith('Unnamed')]

            wb = load_workbook(io.BytesIO(r.content), data_only=True)
            if EXCEL_SHEET not in wb.sheetnames:
                continue
            ws = wb[EXCEL_SHEET]

            status_col_idx = None
            for idx, col in enumerate(selected_cols):
                if str(col).strip().lower() == 'overall status':
                    status_col_idx = idx
                    break

            columns_json = json.dumps([str(c) for c in selected_cols]) if selected_cols else None

            all_cells: list[list] = []
            if len(df) > 0 and selected_cols:
                from openpyxl.utils import get_column_letter
                for row_idx, row in df.iterrows():
                    project_id_val = row[selected_cols[0]]
                    project_id_str = str(project_id_val).strip().lower()
                    if not (
                        pd.notnull(project_id_val)
                        and project_id_str not in ['', 'nan', 'none', '<na>']
                    ):
                        continue

                    cells = []
                    for col_idx, col in enumerate(selected_cols):
                        val = row[col]
                        sval = str(val).strip().lower()

                        if col_idx == 0:
                            if isinstance(val, float) and val.is_integer():
                                cells.append(str(int(val)))
                            elif sval in ['nan', 'none', '<na>', '']:
                                cells.append('')
                            else:
                                cells.append(str(val))
                        elif str(col).strip().lower() in ['plan start', 'plan end']:
                            try:
                                if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                    date_val = pd.to_datetime(val)
                                    if date_val.year == 1970 and date_val.month == 1 and date_val.day == 1:
                                        cells.append('')
                                    else:
                                        cells.append(date_val.date().strftime('%m/%d/%Y'))
                                else:
                                    cells.append('')
                            except Exception:
                                cells.append('')
                        elif str(col).strip().lower() in ['planned %', 'actual %', 'passed %']:
                            try:
                                if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                    percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                    cells.append(f"{percent_val:.2f}%")
                                else:
                                    cells.append('')
                            except Exception:
                                cells.append('')
                        elif col_idx == status_col_idx:
                            excel_col = get_column_letter(col_idx + 1)
                            excel_row = row_idx + 3  # header=1, so first data row is Excel row 3
                            cell_obj = ws[f"{excel_col}{excel_row}"]
                            fill = cell_obj.fill
                            font = cell_obj.font
                            bg_color = None
                            font_color = None

                            cell_value = str(val) if sval not in ['nan', 'none', '<na>', ''] else ''
                            status_val = cell_value.strip().upper()
                            if status_val == 'NOT STARTED':
                                bg_color = None
                            else:
                                if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
                                    bg_color = f"#{fill.fgColor.rgb[2:]}" if fill.fgColor.rgb.startswith('FF') else f"#{fill.fgColor.rgb}"

                            if font and font.color and font.color.type == 'rgb' and font.color.rgb:
                                font_color = f"#{font.color.rgb[2:]}" if font.color.rgb.startswith('FF') else f"#{font.color.rgb}"

                            cells.append({
                                "value": cell_value,
                                "bg_color": bg_color,
                                "font_color": font_color,
                            })
                        else:
                            if sval in ['nan', 'none', '<na>', '']:
                                cells.append('')
                            else:
                                cells.append(str(val))

                    all_cells.append(cells)

            tc_summary_data: list[list[str]] = []
            if 'TC Summary' in wb.sheetnames:
                ws_tc = wb['TC Summary']
                for row_tc in ws_tc.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None and str(cell).strip() != '' for cell in row_tc):
                        tc_summary_data.append([str(cell) if cell is not None else '' for cell in row_tc])

            detailed_tc_headers: list[str] = []
            detailed_tc_data: list[list[str]] = []
            if 'Detailed TC Summary' in wb.sheetnames:
                ws_dtc = wb['Detailed TC Summary']
                all_headers: list[str] = []
                wanted_headers = [
                    'TC No.', 'TC Name', 'Status', 'Start Date', 'End Date',
                    'Total Steps', 'Planned Steps', 'Actual Steps Passed',
                    'Planned %', 'Actual %',
                ]

                def norm(h: str) -> str:
                    return (
                        str(h).strip().lower().replace(' ', '').replace('%', 'percent')
                    )

                header_indices: list[int | None] = []
                for idx2, row_dtc in enumerate(ws_dtc.iter_rows(values_only=True)):
                    if idx2 == 2:
                        all_headers = [str(cell).strip() if cell is not None else '' for cell in row_dtc]
                        normed = [norm(h) for h in all_headers]
                        detailed_tc_headers = [wh for wh in wanted_headers]
                        header_indices = []
                        for wh in wanted_headers:
                            nwh = norm(wh)
                            found = None
                            for j, nh in enumerate(normed):
                                if nwh == nh:
                                    found = j
                                    break
                            header_indices.append(found)
                    elif idx2 > 2 and all_headers and detailed_tc_headers:
                        filtered_row: list[str] = []
                        for col_idx, wh in zip(header_indices, wanted_headers):
                            val = row_dtc[col_idx] if col_idx is not None and col_idx < len(row_dtc) else None
                            wh_norm = wh.strip().lower()
                            if wh_norm in ['planned %', 'actual %']:
                                try:
                                    if val is not None and str(val).strip() != '' and str(val).strip().lower() not in ['nan', 'none', '<na>']:
                                        percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                        filtered_row.append(f"{percent_val:.2f}%")
                                    else:
                                        filtered_row.append('')
                                except Exception:
                                    filtered_row.append('')
                            else:
                                filtered_row.append(str(val) if val is not None else '')

                        if any(cell for cell in filtered_row):
                            detailed_tc_data.append(filtered_row)

            open_defects_data: list[list[str]] = []
            if 'Open Defects' in wb.sheetnames:
                ws_od = wb['Open Defects']
                for row_od in ws_od.iter_rows(values_only=True):
                    open_defects_data.append([str(cell) if cell is not None else '' for cell in row_od])

            with sqlite3.connect(DB_PATH) as conn:
                cur = conn.cursor()
                cur.execute(
                    f'SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data FROM {table_name} WHERE filename = ?',
                    (filename,),
                )
                existing = cur.fetchone()
                new_data = [
                    columns_json,
                    json.dumps(all_cells),
                    json.dumps(tc_summary_data),
                    json.dumps(open_defects_data),
                    json.dumps(detailed_tc_headers),
                    json.dumps(detailed_tc_data),
                ]

                if existing is None:
                    cur.execute(
                        f'INSERT INTO {table_name} (filename, project_name, columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (
                            filename,
                            project_name,
                            columns_json,
                            json.dumps(all_cells),
                            json.dumps(tc_summary_data),
                            json.dumps(open_defects_data),
                            json.dumps(detailed_tc_headers),
                            json.dumps(detailed_tc_data),
                        ),
                    )
                    imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
                else:
                    if list(existing) != new_data:
                        cur.execute(
                            f'UPDATE {table_name} SET columns=?, cells=?, tc_summary=?, open_defects=?, detailed_tc_headers=?, detailed_tc_data=?, project_name=? WHERE filename=?',
                            (
                                columns_json,
                                json.dumps(all_cells),
                                json.dumps(tc_summary_data),
                                json.dumps(open_defects_data),
                                json.dumps(detailed_tc_headers),
                                json.dumps(detailed_tc_data),
                                project_name,
                                filename,
                            ),
                        )
                        imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
        except Exception:
            traceback.print_exc()
            continue

    total = sum(imported_counts.values())

    # After import, enforce "latest per (PID, Project Name)" dedup on the target tables.
    # This prevents older snapshots from accumulating in DB over time.
    dedup_deleted: dict[str, int] = {}
    if tags_filter:
        for tag in tags_filter:
            table_name = _table_for_tag(tag)
            try:
                dedup_deleted[table_name] = _deduplicate_table_latest_by_pid_project(table_name)
            except Exception:
                traceback.print_exc()

    dedup_msg = ''
    if dedup_deleted:
        parts = [f"{tbl}:{cnt}" for tbl, cnt in dedup_deleted.items() if cnt]
        if parts:
            dedup_msg = f" Dedup deleted {', '.join(parts)} old rows."
    if tags_filter:
        tags_str = ','.join(sorted(tags_filter))
        return (
            f"Imported {total} files for tags [{tags_str}]. "
            f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}",
            200,
        )

    return (
        f"Imported {total} files. "
        f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}",
        200,
    )


# --- ADMIN: Import from GitHub to DB (should be protected in production) ---
@app.route('/admin/import', methods=['POST'])
def import_from_github():
    try:
        return _import_from_github_routed(env_prefix='ETE', tags_filter={'ETE'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)


@app.route('/admin/import_pvt', methods=['POST'])
def import_from_github_pvt():
    try:
        return _import_from_github_routed(env_prefix='PVT', tags_filter={'PVT'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)


@app.route('/admin/import_billing', methods=['POST'])
def import_from_github_billing():
    try:
        return _import_from_github_routed(env_prefix='BILLING', tags_filter={'ETE_Billing'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)

# Main dashboard route
@app.route('/')
def main_dashboard():
    html_path = os.path.join(os.path.dirname(__file__), 'main_dashboard.html')
    return send_file(html_path)


# ETE Project Dashboard route (now fetches from DB)
@app.route('/ete')
def ete_dashboard():
    # Read dashboard-ready rows from dashboard_data table
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data, project_name FROM dashboard_data')
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    # Reconstruct columns and rows
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        tc_summary = json.loads(rec[2])
        open_defects = json.loads(rec[3])
        detailed_tc_headers = json.loads(rec[4])
        detailed_tc_data = json.loads(rec[5])
        project_name = rec[6] if len(rec) > 6 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            # Insert Project Name as the second column if not present
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            # Insert project_name as the second cell if not present
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            # Deduplication key: (Project ID, Project Name)
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                dedup_dict[key] = {
                    "cells": cells,
                    "tc_summary": tc_summary,
                    "open_defects": open_defects,
                    "detailed_tc_headers": detailed_tc_headers,
                    "detailed_tc_data": detailed_tc_data
                }
            except Exception:
                # fallback: just append
                all_rows.append({
                    "cells": cells,
                    "tc_summary": tc_summary,
                    "open_defects": open_defects,
                    "detailed_tc_headers": detailed_tc_headers,
                    "detailed_tc_data": detailed_tc_data
                })
    # Use deduplicated rows if any
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    # Find status_col_idx
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    # Default sort by Plan Start descending if column exists
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='ETE Project Dashboard - L1 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )

if __name__ == '__main__':
    try:
        print('Starting Flask app on http://localhost:5000 ...')
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
    except KeyboardInterrupt:
        print('\nServer stopped by user (Ctrl+C)')


