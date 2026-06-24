import io
import json
import os
import re
import sqlite3
import traceback
from datetime import datetime

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template_string, send_file, request
import threading
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

load_dotenv()

from db_utils import DB_PATH

app = Flask(__name__)

# --- PVT Project Dashboard route ---
@app.route('/pvt')
def pvt_dashboard():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Filter for PVT project data (assumes project_name or similar field distinguishes)
    cursor.execute("SELECT columns, cells, project_name FROM dashboard_data_pvt ORDER BY id DESC")
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        project_name = rec[2] if len(rec) > 2 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                # Records are ordered newest-first; keep the first seen for each (PID, Project Name)
                if key not in dedup_dict:
                    dedup_dict[key] = {
                        "cells": cells
                    }
            except Exception:
                all_rows.append({
                    "cells": cells
                })
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='PVT Project Dashboard - L2 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )


# --- ETE Billing Dashboard route ---
@app.route('/ete-billing')
def ete_billing_dashboard():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Filter for ETE Billing project data (assumes project_name or similar field distinguishes)
    cursor.execute("SELECT columns, cells, project_name FROM dashboard_data_billing ORDER BY id DESC")
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        project_name = rec[2] if len(rec) > 2 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                # Records are ordered newest-first; keep the first seen for each (PID, Project Name)
                if key not in dedup_dict:
                    dedup_dict[key] = {
                        "cells": cells
                    }
            except Exception:
                all_rows.append({
                    "cells": cells
                })
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='ETE Billing Project Dashboard - L2 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )
dashboard_html = r'''
<!DOCTYPE html>
<html lang="en">
<head>
    <!-- Font Awesome for Home Icon -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css">
    <!-- Google Fonts: Poppins -->
    <link href="https://fonts.googleapis.com/css?family=Poppins:600,700&display=swap" rel="stylesheet">
    <meta charset="UTF-8">
    <title>Project Dashboard</title>
    <style>
        /* Reduce font size for Project ID column */
        #mainTable td:first-child {
                font-size: 13px !important;
                text-align: center !important;
                vertical-align: middle !important;
                font-family: 'Times New Roman', Times, serif !important;
                font-weight: normal !important;
                text-transform: none !important;
                letter-spacing: normal !important;
                min-width: 75px;
                width: 75px;
                max-width: 75px;
                word-break: break-word;
                white-space: normal !important;
                padding-left: 0 !important;
            }
        }
    body, table, th, td, .container, .no-data, button, div, span {
        font-family: 'Times New Roman', Times, serif !important;
        font-size: 16px !important;
    }
    .container {
        max-width: 900px;
        margin: 40px auto;
        background: linear-gradient(135deg, rgba(253,246,238,0.5) 0%, rgba(243,247,250,0.5) 100%);
        padding: 2rem;
        border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.11);
        text-align: center;
        font-family: 'Times New Roman', Times, serif;
        backdrop-filter: blur(8px);
        background: rgba(255,255,255,0.55);
        border: 1px solid rgba(255,255,255,0.18);
    }
    h2 {
        text-align: center;
        font-family: 'Poppins', 'Times New Roman', Times, serif !important;
        font-size: 28px !important;
        font-weight: bold;
    }
    .table-responsive {
        width: 100%;
        margin-top: 2rem;
        /* overflow: visible;  -- ensure dropdowns are not clipped */
    }
    table, .open-defects-table {
        width: auto;
        margin-left: -18px;
        margin-right: auto;
        border-collapse: separate;
        border-spacing: 0;
        border: 1.5px solid #e0c9a6;
        background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%);
        border-radius: 12px;
        box-shadow: 0 4px 16px rgba(180,150,100,0.10), 0 1.5px 4px rgba(180,150,100,0.08);
        overflow: visible;
        table-layout: auto;
    }
    th, td {
        font-size: 13px !important;
        padding: 4px 6px;
        white-space: normal !important;
        word-break: break-word;
        max-width: 180px;
        font-family: 'Times New Roman', Times, serif !important;
        font-weight: normal !important;
        text-transform: none !important;
        overflow: visible !important; /* ensure dropdowns are not clipped */
    }
    th, td {
        border: 1px solid #e0c9a6;
        padding: 8px;
        text-align: center;
        vertical-align: middle;
    }
    th {
        background: linear-gradient(135deg, #f5e9da 0%, #e9d8c3 100%) !important;
        color: #7a5c2e !important;
        font-weight: bold;
        font-size: 17px;
        border-bottom: 3px solid #e0c9a6;
    }
    /* Zebra striping for all tables except header row */
    table tr:not(:first-child):nth-of-type(even), .open-defects-table tr:not(:first-child):nth-of-type(even) {
        background: #f9f3e7 !important;
    }
    table tr:not(:first-child):nth-of-type(odd), .open-defects-table tr:not(:first-child):nth-of-type(odd) {
        background: #fdf6ee !important;
    }
    .no-data { color: #b00; margin: 1rem 0; }
    table tr:not(:first-child):hover, .open-defects-table tr:not(:first-child):hover {
        background-color: #f3e7d1 !important;
        transition: background 0.2s;
    }
    /* Status color classes for Detailed TC Summary */
    .status-blocked { color: #d32f2f !important; font-weight: bold; }
    .status-failed { color: #d32f2f !important; font-weight: bold; }
    .status-in-progress { color: #388e3c !important; font-weight: bold; }
    .status-completed { color: #1976d2 !important; font-weight: bold; }
    .status-execution-completed { color: #1976d2 !important; font-weight: bold; }
    .status-execution-complete-uploaded-for-review { color: #1976d2 !important; font-weight: bold; }
    .status-on-hold { color: #ff9800 !important; font-weight: bold; }
    .status-not-started { color: #757575 !important; font-weight: bold; }
    .status-deferred { color: #757575 !important; font-weight: bold; }
    .status-descoped { color: #757575 !important; font-weight: bold; }
    .status-default { color: #222 !important; font-weight: bold; }
    </style>
</head>
<body style="position:relative; min-height:100vh; background: #f7f3ea;">
<!-- IBM and AT&T Logos -->
<img src="/static/ibm.png" alt="IBM Logo" style="position:fixed; top:24px; left:32px; height:44px; z-index:10; background:transparent; border-radius:8px; padding:4px 10px; box-shadow:0 2px 8px rgba(0,0,0,0.07);">
<img src="/static/att.png" alt="AT&T Logo" style="position:fixed; top:24px; right:32px; height:44px; z-index:10; background:transparent; border-radius:8px; padding:4px 10px; box-shadow:0 2px 8px rgba(0,0,0,0.07);">
<!-- Soft beige wave SVG background -->
<div style="position:fixed; z-index:-1; inset:0; width:100vw; height:100vh; pointer-events:none;">
    <svg width="100%" height="100%" viewBox="0 0 1920 1080" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg">
        <defs>
            <linearGradient id="beigeGrad" x1="0" y1="0" x2="0" y2="1">
                <stop offset="0%" stop-color="#ede3ce"/>
                <stop offset="100%" stop-color="#f7f3ea"/>
            </linearGradient>
        </defs>
        <rect width="1920" height="1080" fill="url(#beigeGrad)"/>
        <path d="M0,200 Q600,350 1920,120 Q1920,0 0,0 Z" fill="#e8dbc0" fill-opacity="0.45"/>
        <path d="M0,600 Q900,900 1920,400 Q1920,0 0,0 Z" fill="#e2d2b6" fill-opacity="0.32"/>
        <path d="M0,900 Q1200,1200 1920,700 Q1920,0 0,0 Z" fill="#e6dcc5" fill-opacity="0.22"/>
    </svg>
</div>
<div class="container" style="position:relative;">
    <a href="/" style="position:absolute; top:22px; right:28px; z-index:2; text-decoration:none;">
        <i class="fa-solid fa-house home-icon-ete"></i>
    </a>
    <style>
    .home-icon-ete {
        font-size: 1.3em;
        color: #7a5c2e;
        background: rgba(255,255,255,0.45);
        border-radius: 12px;
        padding: 8px 12px;
        box-shadow: 0 4px 16px rgba(180,150,100,0.10);
        backdrop-filter: blur(4px);
        transition: background 0.18s, color 0.18s;
        cursor: pointer;
    }
    .home-icon-ete:hover {
        background: rgba(245,233,218,0.85);
        color: #5a3d13;
    }
    </style>
<div style="margin-bottom: 1.2em; position: relative;">
    <h2 style="margin: 0; text-align: center; font-size: 1.5em; font-family: 'Times New Roman', Times, serif !important; font-weight: bold;">{{ dashboard_title|default('ETE Project Dashboard - L2 Report') }}</h2>
</div>
{% if errors %}
    {% for error in errors %}
        <div style="color:red;">{{ error }}</div>
    {% endfor %}
{% endif %}
{% if not rows or rows|length == 0 %}
    <div class="no-data">No data found in any file.</div>
{% else %}
    <div class="table-responsive">
    <!-- Filter box above the table, aligned with Overall Status column -->
    <div style="display: flex; justify-content: flex-end; margin-bottom: 8px;">
        <div style="position: relative; display: inline-block; margin-left: 24px;">
            <input id="globalFilterBox" type="text" placeholder="Filter all columns..." style="padding: 6px 12px; border: 1.2px solid #e0c9a6; border-radius: 6px; font-size: 14px; font-family: 'Times New Roman', Times, serif; width: 120px; background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%); box-shadow: 0 2px 6px rgba(180,150,100,0.07); outline: none; transition: border 0.18s, box-shadow 0.18s;" oninput="document.getElementById('clearFilterBox').style.display = this.value ? 'block' : 'none';" />
            <span id="clearFilterBox" style="display:none; position: absolute; right: 8px; top: 50%; transform: translateY(-50%); cursor: pointer; color: #bfa16a; font-size: 16px; user-select: none;" onclick="var f=document.getElementById('globalFilterBox');f.value='';f.dispatchEvent(new Event('input'));this.style.display='none';">&#10005;</span>
        </div>
    </div>
<table id="mainTable">
    <thead>
        <tr>
            {% for col in columns %}
            <th style="position: relative; white-space:nowrap; padding: 8px 10px;{% if col|string|trim|lower == 'overall status' %} min-width:130px;{% elif col|string|trim|lower == 'plan end' %} min-width:60px;{% endif %}" class="{% if col|string|trim|lower == 'overall status' %}status-col{% endif %}">
                <div style="display:flex; justify-content:center; align-items:center; gap:2px; margin-top:2px;">
                    <span class="sort-icons" style="cursor:pointer;display:inline-flex;flex-direction:column;align-items:center;gap:1px; margin-right:2px;" onclick="sortTable({{ loop.index0 }})">
                        <svg width="12" height="8" viewBox="0 0 16 10" style="display:block;">
                            <polygon points="8,2 3,8 13,8" fill="#222"/>
                        </svg>
                        <svg width="12" height="8" viewBox="0 0 16 10" style="display:block;">
                            <polygon points="8,8 3,2 13,2" fill="#222"/>
                        </svg>
                    </span>
                    <div style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis; font-weight:600; font-size:{% if loop.index0 == 0 %}12px{% else %}1em{% endif %};">{{ col }}</div>
                </div>
                <div class="column-filter-dropdown" data-col="{{ loop.index0 }}" style="display:none; position: absolute; top: 100%; left: 0; background: none; border: none; border-radius: 9px; box-shadow: none; padding: 6px 10px; z-index: 10; min-width: 110px;">
                    <input type="text" class="column-filter-input modern-filter-input" placeholder="Type to filter..." style="width: 100%; padding: 5px 9px; border: 1.2px solid #e0c9a6; border-radius: 6px; font-size: 13px; font-family: 'Times New Roman', Times, serif; background: linear-gradient(135deg, #fdf6ee 0%, #f3f7fa 100%); box-shadow: 0 2px 6px rgba(180,150,100,0.07); outline: none; transition: border 0.18s, box-shadow 0.18s;">
                </div>
            </th>
            {% endfor %}
        </tr>
    </thead>
    <tbody>
        {% for row in rows %}
        <tr>
            {% for cell in row.cells %}
                {% set cell_idx = loop.index0 %}
                {% set display_cell = cell if cell|string|trim != '' else 'N/A' %}
                {% if status_col_idx is not none and cell_idx == status_col_idx and cell is mapping %}
                    <td class="status-col" style="background: transparent; color: {{ cell.font_color if cell.font_color else 'black' }}; text-align:center; vertical-align:middle;">
                        <span style="
                            display: inline-block;
                            border-radius: 16px;
                            padding: 6px 18px;
                            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
                            border: 1.5px solid #fff;
                            min-width: 90px;
                            background-color: {{ cell.bg_color if cell.bg_color and cell.bg_color|lower not in ['#1976d2', '#b7d7f7', '#e3f0fc', '#a4c8ea', '#007bff', '#154360'] else '' }};
                            color: #fff !important;
                            font-family: 'Times New Roman', Times, serif !important;
                            font-size: 13px !important;
                            font-weight: bold !important;
                            text-transform: none !important;
                            letter-spacing: normal !important;
                        ">
                            {{ cell.value if cell.value|string|trim != '' else 'N/A' }}
                        </span>
                    </td>
                {% else %}
                    <td>{{ display_cell }}</td>
                {% endif %}
            {% endfor %}
        </tr>
        {% endfor %}
    </tbody>
    </table>
    </div>
    <script>
                // --- Sort Table Function ---
                function sortTable(colIdx) {
                    var table = document.getElementById('mainTable');
                    var tbody = table.querySelector('tbody');
                    if (!tbody) return;
                    var rows = Array.from(tbody.querySelectorAll('tr'));
                    // Track sort direction
                    if (!table._sortState) table._sortState = {};
                    var state = table._sortState;
                    state[colIdx] = state[colIdx] === 'asc' ? 'desc' : 'asc';
                    var dir = state[colIdx];
                    rows.sort(function(a, b) {
                        var cellA = a.querySelectorAll('td')[colIdx];
                        var cellB = b.querySelectorAll('td')[colIdx];
                        var valA = cellA ? cellA.textContent.trim() : '';
                        var valB = cellB ? cellB.textContent.trim() : '';
                        // Try to compare as numbers, fallback to string
                        var numA = parseFloat(valA.replace(/[^0-9.\-]/g, ''));
                        var numB = parseFloat(valB.replace(/[^0-9.\-]/g, ''));
                        if (!isNaN(numA) && !isNaN(numB)) {
                            return dir === 'asc' ? numA - numB : numB - numA;
                        } else {
                            valA = valA.toLowerCase();
                            valB = valB.toLowerCase();
                            if (valA < valB) return dir === 'asc' ? -1 : 1;
                            if (valA > valB) return dir === 'asc' ? 1 : -1;
                            return 0;
                        }
                    });
                    rows.forEach(function(tr) { tbody.removeChild(tr); });
                    rows.forEach(function(tr) { tbody.appendChild(tr); });
                }
        // --- Filter Icon Highlighting ---
        function updateFilterIconHighlight() {
            // Main table
            document.querySelectorAll('.column-filter-dropdown').forEach(function(dropdown) {
                var colIdx = dropdown.getAttribute('data-col');
                var input = dropdown.querySelector('input');
                var icon = dropdown.parentElement.querySelector('.filter-svg');
                if (input && icon) {
                    if (input.value.trim() !== '') {
                        icon.querySelector('path').setAttribute('stroke', '#333');
                    } else {
                        icon.querySelector('path').setAttribute('stroke', '#888');
                    }
                }
            });
        }
    // Column Filter Dropdown Logic for Main Table
    function toggleColumnFilter(icon, colIdx, event) {
        // Hide all other filter dropdowns
        document.querySelectorAll('.column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
        // Find the dropdown for this column (relative to the <th>)
        var th = icon.closest('th');
        var dropdown = th ? th.querySelector('.column-filter-dropdown') : null;
        if (dropdown) {
            // Show the dropdown
            dropdown.style.display = 'block';
            var input = dropdown.querySelector('input');
            if (input) { input.focus(); }
        }
        // Prevent click from closing the dropdown
        if (event) event.stopPropagation();
    }
    document.addEventListener('click', function() {
        document.querySelectorAll('.column-filter-dropdown').forEach(function(el) { el.style.display = 'none'; });
    });
    document.querySelectorAll('.column-filter-input').forEach(function(input) {
        input.addEventListener('input', function() {
            var colIdx = parseInt(input.parentElement.getAttribute('data-col'));
            var filter = input.value.toLowerCase();
            var table = document.getElementById('mainTable');
            var trs = table.querySelectorAll('tbody > tr');
            trs.forEach(function(tr) {
                if (!tr.querySelectorAll('td').length) return;
                var tds = tr.querySelectorAll('td');
                var cell = tds[colIdx];
                if (cell) {
                    var text = cell.textContent.toLowerCase();
                    tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                }
            });
            updateFilterIconHighlight();
        });
    });

    // Main Table Filter
    document.addEventListener('DOMContentLoaded', function() {
                updateFilterIconHighlight();
                // Also update on page load in case filters are pre-filled
                document.querySelectorAll('.column-filter-input').forEach(function(input) {
                    input.addEventListener('change', updateFilterIconHighlight);
                });
        var mainFilter = document.getElementById('mainTableFilter');
        if (mainFilter) {
            mainFilter.addEventListener('input', function() {
                var filter = mainFilter.value.toLowerCase();
                var table = document.getElementById('mainTable');
                var trs = table.querySelectorAll('tbody > tr');
                trs.forEach(function(tr) {
                    // Only filter main rows, not details rows
                    if (tr.querySelectorAll('td').length && !tr.id.startsWith('details-')) {
                        var text = tr.textContent.toLowerCase();
                        tr.style.display = text.indexOf(filter) > -1 ? '' : 'none';
                    }
                });
            });
        }
    });
    // Global filter for all columns
    var globalFilterBox = document.getElementById('globalFilterBox');
    if (globalFilterBox) {
        globalFilterBox.addEventListener('input', function() {
            var filter = globalFilterBox.value.toLowerCase();
            var table = document.getElementById('mainTable');
            if (!table) return;
            var trs = table.querySelectorAll('tbody > tr');
            trs.forEach(function(tr) {
                var text = tr.textContent.toLowerCase();
                if (text.indexOf(filter) > -1) {
                    tr.style.display = '';
                } else {
                    tr.style.display = 'none';
                }
            });
        });
    }
    // ...existing code...
    </script>
{% endif %}
</div>
</body>
</html>
'''

# --- CONFIG ---
EXCEL_SHEET = 'Overall Status'  # Used for import only


def _ensure_dashboard_table(table_name: str) -> None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {table_name} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT UNIQUE,
                project_name TEXT,
                columns TEXT,
                cells TEXT,
                tc_summary TEXT,
                open_defects TEXT,
                detailed_tc_headers TEXT,
                detailed_tc_data TEXT
            )'''
        )


def _parse_dashboard_filename(filename: str) -> tuple[str, str, str, int, int] | None:
    """Parse PID dashboard filename.

    Expected patterns:
      PID_<pid>_<project_name>_ETE_Status_as_of_<MM>_<DD>.xlsx
      PID_<pid>_<project_name>_PVT_Status_as_of_<MM>_<DD>.xlsx
      PID_<pid>_<project_name>_ETE_Billing_Status_as_of_<MM>_<DD>.xlsx
    """
    pattern = re.compile(
        r'^PID_(\d+)_(.+?)_(ETE_Billing|ETE|PVT)_Status_as_of_(\d{2})_(\d{2})',
        re.IGNORECASE,
    )
    m = pattern.search(filename or '')
    if not m:
        return None
    pid = m.group(1)
    project_name_raw = m.group(2)
    tag = m.group(3)
    mm = int(m.group(4))
    dd = int(m.group(5))
    project_name = project_name_raw.replace('_', ' ').strip()
    # Normalize tag casing
    tag_norm = tag.strip()
    if tag_norm.lower() == 'ete_billing':
        tag_norm = 'ETE_Billing'
    else:
        tag_norm = tag_norm.upper()
    return pid, project_name, tag_norm, mm, dd


def _table_for_tag(tag: str) -> str:
    tag_norm = (tag or '').strip()
    if tag_norm == 'PVT':
        return 'dashboard_data_pvt'
    if tag_norm in ('ETE_Billing', 'ETE BILLING'):
        return 'dashboard_data_billing'
    return 'dashboard_data'


def _deduplicate_table_latest_by_pid_project(table_name: str) -> int:
    """Keep only the latest (MM,DD) record per (PID, project_name) in the given table.

    Latest-ness is determined from the filename pattern `_Status_as_of_MM_DD`.
    Returns number of deleted rows.
    """
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id, filename, project_name FROM {table_name}")
        rows = cur.fetchall()

        latest: dict[tuple[str, str], tuple[int, int, int]] = {}
        # key -> (id, mm, dd)
        for row_id, filename, project_name in rows:
            parsed = _parse_dashboard_filename(filename or '')
            if not parsed:
                continue
            pid, pname, _tag, mm, dd = parsed
            key = (str(pid).strip(), str(pname or project_name or '').strip())

            if key not in latest or (mm, dd) > (latest[key][1], latest[key][2]):
                latest[key] = (int(row_id), int(mm), int(dd))

        keep_ids = {v[0] for v in latest.values()}
        if not keep_ids:
            return 0

        placeholders = ','.join(['?'] * len(keep_ids))
        cur.execute(
            f"DELETE FROM {table_name} WHERE id NOT IN ({placeholders})",
            tuple(keep_ids),
        )
        deleted = cur.rowcount if cur.rowcount is not None else 0
        conn.commit()
        return deleted


def _import_from_github_routed(*, env_prefix: str, tags_filter: set[str] | None = None) -> tuple[str, int]:
    import requests
    from openpyxl import load_workbook

    token = os.environ.get(f'{env_prefix}_GITHUB_TOKEN') or os.environ.get('GITHUB_TOKEN')
    user = os.environ.get(f'{env_prefix}_GITHUB_USER') or os.environ.get('GITHUB_USER')
    repo = os.environ.get(f'{env_prefix}_GITHUB_REPO') or os.environ.get('GITHUB_REPO')

    if not (token and user and repo):
        return (
            f'Missing GitHub credentials. Set {env_prefix}_GITHUB_TOKEN/{env_prefix}_GITHUB_USER/{env_prefix}_GITHUB_REPO (or the generic GITHUB_TOKEN/GITHUB_USER/GITHUB_REPO).',
            403,
        )

    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://api.github.com/repos/{user}/{repo}/contents"

    r = requests.get(url, headers=headers)
    r.raise_for_status()
    files = r.json()
    excel_files = [f for f in files if f.get('name', '').lower().endswith(('.xlsx', '.xls'))]

    latest_file_map: dict[tuple[str, str, str], tuple[tuple[int, int], dict]] = {}
    for f in excel_files:
        filename = f.get('name', '')
        parsed = _parse_dashboard_filename(filename)
        if not parsed:
            continue
        pid, project_name, tag, mm, dd = parsed
        if tags_filter and tag not in tags_filter:
            continue
        key = (pid, project_name, tag)
        if key not in latest_file_map or (mm, dd) > latest_file_map[key][0]:
            latest_file_map[key] = ((mm, dd), f)

    latest_files = [v[1] for v in latest_file_map.values()]

    imported_counts: dict[str, int] = {
        'dashboard_data': 0,
        'dashboard_data_pvt': 0,
        'dashboard_data_billing': 0,
    }

    for file_info in latest_files:
        download_url = file_info.get('download_url')
        filename = file_info.get('name', '')
        if not download_url or not filename:
            continue

        r = requests.get(download_url, headers=headers)
        if r.status_code != 200:
            continue

        try:
            parsed = _parse_dashboard_filename(filename)
            if not parsed:
                continue
            pid, project_name, tag, mm, dd = parsed
            if tags_filter and tag not in tags_filter:
                continue

            table_name = _table_for_tag(tag)
            _ensure_dashboard_table(table_name)

            df = pd.read_excel(io.BytesIO(r.content), sheet_name=EXCEL_SHEET, header=1)
            selected_cols = [col for col in df.columns if not str(col).startswith('Unnamed')]

            wb = load_workbook(io.BytesIO(r.content), data_only=True)
            if EXCEL_SHEET not in wb.sheetnames:
                continue
            ws = wb[EXCEL_SHEET]

            status_col_idx = None
            for idx, col in enumerate(selected_cols):
                if str(col).strip().lower() == 'overall status':
                    status_col_idx = idx
                    break

            columns_json = json.dumps([str(c) for c in selected_cols]) if selected_cols else None

            all_cells: list[list] = []
            if len(df) > 0 and selected_cols:
                from openpyxl.utils import get_column_letter
                for row_idx, row in df.iterrows():
                    project_id_val = row[selected_cols[0]]
                    project_id_str = str(project_id_val).strip().lower()
                    if not (
                        pd.notnull(project_id_val)
                        and project_id_str not in ['', 'nan', 'none', '<na>']
                    ):
                        continue

                    cells = []
                    for col_idx, col in enumerate(selected_cols):
                        val = row[col]
                        sval = str(val).strip().lower()

                        if col_idx == 0:
                            if isinstance(val, float) and val.is_integer():
                                cells.append(str(int(val)))
                            elif sval in ['nan', 'none', '<na>', '']:
                                cells.append('')
                            else:
                                cells.append(str(val))
                        elif str(col).strip().lower() in ['plan start', 'plan end']:
                            try:
                                if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                    date_val = pd.to_datetime(val)
                                    if date_val.year == 1970 and date_val.month == 1 and date_val.day == 1:
                                        cells.append('')
                                    else:
                                        cells.append(date_val.date().strftime('%m/%d/%Y'))
                                else:
                                    cells.append('')
                            except Exception:
                                cells.append('')
                        elif str(col).strip().lower() in ['planned %', 'actual %', 'passed %']:
                            try:
                                if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                    percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                    cells.append(f"{percent_val:.2f}%")
                                else:
                                    cells.append('')
                            except Exception:
                                cells.append('')
                        elif col_idx == status_col_idx:
                            excel_col = get_column_letter(col_idx + 1)
                            excel_row = row_idx + 3  # header=1, so first data row is Excel row 3
                            cell_obj = ws[f"{excel_col}{excel_row}"]
                            fill = cell_obj.fill
                            font = cell_obj.font
                            bg_color = None
                            font_color = None

                            cell_value = str(val) if sval not in ['nan', 'none', '<na>', ''] else ''
                            status_val = cell_value.strip().upper()
                            if status_val == 'NOT STARTED':
                                bg_color = None
                            else:
                                if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
                                    bg_color = f"#{fill.fgColor.rgb[2:]}" if fill.fgColor.rgb.startswith('FF') else f"#{fill.fgColor.rgb}"

                            if font and font.color and font.color.type == 'rgb' and font.color.rgb:
                                font_color = f"#{font.color.rgb[2:]}" if font.color.rgb.startswith('FF') else f"#{font.color.rgb}"

                            cells.append({
                                "value": cell_value,
                                "bg_color": bg_color,
                                "font_color": font_color,
                            })
                        else:
                            if sval in ['nan', 'none', '<na>', '']:
                                cells.append('')
                            else:
                                cells.append(str(val))

                    all_cells.append(cells)

            tc_summary_data: list[list[str]] = []
            if 'TC Summary' in wb.sheetnames:
                ws_tc = wb['TC Summary']
                for row_tc in ws_tc.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None and str(cell).strip() != '' for cell in row_tc):
                        tc_summary_data.append([str(cell) if cell is not None else '' for cell in row_tc])

            detailed_tc_headers: list[str] = []
            detailed_tc_data: list[list[str]] = []
            if 'Detailed TC Summary' in wb.sheetnames:
                ws_dtc = wb['Detailed TC Summary']
                all_headers: list[str] = []
                wanted_headers = [
                    'TC No.', 'TC Name', 'Status', 'Start Date', 'End Date',
                    'Total Steps', 'Planned Steps', 'Actual Steps Passed',
                    'Planned %', 'Actual %',
                ]

                def norm(h: str) -> str:
                    return (
                        str(h).strip().lower().replace(' ', '').replace('%', 'percent')
                    )

                header_indices: list[int | None] = []
                for idx2, row_dtc in enumerate(ws_dtc.iter_rows(values_only=True)):
                    if idx2 == 2:
                        all_headers = [str(cell).strip() if cell is not None else '' for cell in row_dtc]
                        normed = [norm(h) for h in all_headers]
                        detailed_tc_headers = [wh for wh in wanted_headers]
                        header_indices = []
                        for wh in wanted_headers:
                            nwh = norm(wh)
                            found = None
                            for j, nh in enumerate(normed):
                                if nwh == nh:
                                    found = j
                                    break
                            header_indices.append(found)
                    elif idx2 > 2 and all_headers and detailed_tc_headers:
                        filtered_row: list[str] = []
                        for col_idx, wh in zip(header_indices, wanted_headers):
                            val = row_dtc[col_idx] if col_idx is not None and col_idx < len(row_dtc) else None
                            wh_norm = wh.strip().lower()
                            if wh_norm in ['planned %', 'actual %']:
                                try:
                                    if val is not None and str(val).strip() != '' and str(val).strip().lower() not in ['nan', 'none', '<na>']:
                                        percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                        filtered_row.append(f"{percent_val:.2f}%")
                                    else:
                                        filtered_row.append('')
                                except Exception:
                                    filtered_row.append('')
                            else:
                                filtered_row.append(str(val) if val is not None else '')

                        if any(cell for cell in filtered_row):
                            detailed_tc_data.append(filtered_row)

            open_defects_data: list[list[str]] = []
            if 'Open Defects' in wb.sheetnames:
                ws_od = wb['Open Defects']
                for row_od in ws_od.iter_rows(values_only=True):
                    open_defects_data.append([str(cell) if cell is not None else '' for cell in row_od])

            with sqlite3.connect(DB_PATH) as conn:
                cur = conn.cursor()
                cur.execute(
                    f'SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data FROM {table_name} WHERE filename = ?',
                    (filename,),
                )
                existing = cur.fetchone()
                new_data = [
                    columns_json,
                    json.dumps(all_cells),
                    json.dumps(tc_summary_data),
                    json.dumps(open_defects_data),
                    json.dumps(detailed_tc_headers),
                    json.dumps(detailed_tc_data),
                ]

                if existing is None:
                    cur.execute(
                        f'INSERT INTO {table_name} (filename, project_name, columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (
                            filename,
                            project_name,
                            columns_json,
                            json.dumps(all_cells),
                            json.dumps(tc_summary_data),
                            json.dumps(open_defects_data),
                            json.dumps(detailed_tc_headers),
                            json.dumps(detailed_tc_data),
                        ),
                    )
                    imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
                else:
                    if list(existing) != new_data:
                        cur.execute(
                            f'UPDATE {table_name} SET columns=?, cells=?, tc_summary=?, open_defects=?, detailed_tc_headers=?, detailed_tc_data=?, project_name=? WHERE filename=?',
                            (
                                columns_json,
                                json.dumps(all_cells),
                                json.dumps(tc_summary_data),
                                json.dumps(open_defects_data),
                                json.dumps(detailed_tc_headers),
                                json.dumps(detailed_tc_data),
                                project_name,
                                filename,
                            ),
                        )
                        imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
        except Exception:
            traceback.print_exc()
            continue

    total = sum(imported_counts.values())

    # After import, enforce "latest per (PID, Project Name)" dedup on the target tables.
    # This prevents older snapshots from accumulating in DB over time.
    dedup_deleted: dict[str, int] = {}
    if tags_filter:
        for tag in tags_filter:
            table_name = _table_for_tag(tag)
            try:
                dedup_deleted[table_name] = _deduplicate_table_latest_by_pid_project(table_name)
            except Exception:
                traceback.print_exc()

    dedup_msg = ''
    if dedup_deleted:
        parts = [f"{tbl}:{cnt}" for tbl, cnt in dedup_deleted.items() if cnt]
        if parts:
            dedup_msg = f" Dedup deleted {', '.join(parts)} old rows."
    if tags_filter:
        tags_str = ','.join(sorted(tags_filter))
        return (
            f"Imported {total} files for tags [{tags_str}]. "
            f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}",
            200,
        )

    return (
        f"Imported {total} files. "
        f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}",
        200,
    )


def _process_excel_bytes(filename: str, content: bytes, tags_filter: set[str] | None, imported_counts: dict) -> None:
    """Process a single Excel file given its bytes and insert/update DB tables."""
    from openpyxl import load_workbook

    try:
        parsed = _parse_dashboard_filename(filename)
        if not parsed:
            return
        pid, project_name, tag, mm, dd = parsed
        if tags_filter and tag not in tags_filter:
            return

        table_name = _table_for_tag(tag)
        _ensure_dashboard_table(table_name)

        df = pd.read_excel(io.BytesIO(content), sheet_name=EXCEL_SHEET, header=1)
        selected_cols = [col for col in df.columns if not str(col).startswith('Unnamed')]

        wb = load_workbook(io.BytesIO(content), data_only=True)
        if EXCEL_SHEET not in wb.sheetnames:
            return
        ws = wb[EXCEL_SHEET]

        status_col_idx = None
        for idx, col in enumerate(selected_cols):
            if str(col).strip().lower() == 'overall status':
                status_col_idx = idx
                break

        columns_json = json.dumps([str(c) for c in selected_cols]) if selected_cols else None

        all_cells: list[list] = []
        if len(df) > 0 and selected_cols:
            from openpyxl.utils import get_column_letter
            for row_idx, row in df.iterrows():
                project_id_val = row[selected_cols[0]]
                project_id_str = str(project_id_val).strip().lower()
                if not (
                    pd.notnull(project_id_val)
                    and project_id_str not in ['', 'nan', 'none', '<na>']
                ):
                    continue

                cells = []
                for col_idx, col in enumerate(selected_cols):
                    val = row[col]
                    sval = str(val).strip().lower()

                    if col_idx == 0:
                        if isinstance(val, float) and val.is_integer():
                            cells.append(str(int(val)))
                        elif sval in ['nan', 'none', '<na>', '']:
                            cells.append('')
                        else:
                            cells.append(str(val))
                    elif str(col).strip().lower() in ['plan start', 'plan end']:
                        try:
                            if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                date_val = pd.to_datetime(val)
                                if date_val.year == 1970 and date_val.month == 1 and date_val.day == 1:
                                    cells.append('')
                                else:
                                    cells.append(date_val.date().strftime('%m/%d/%Y'))
                            else:
                                cells.append('')
                        except Exception:
                            cells.append('')
                    elif str(col).strip().lower() in ['planned %', 'actual %', 'passed %']:
                        try:
                            if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                cells.append(f"{percent_val:.2f}%")
                            else:
                                cells.append('')
                        except Exception:
                            cells.append('')
                    elif col_idx == status_col_idx:
                        excel_col = get_column_letter(col_idx + 1)
                        excel_row = row_idx + 3  # header=1, so first data row is Excel row 3
                        cell_obj = ws[f"{excel_col}{excel_row}"]
                        fill = cell_obj.fill
                        font = cell_obj.font
                        bg_color = None
                        font_color = None

                        cell_value = str(val) if sval not in ['nan', 'none', '<na>', ''] else ''
                        status_val = cell_value.strip().upper()
                        if status_val == 'NOT STARTED':
                            bg_color = None
                        else:
                            if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
                                bg_color = f"#{fill.fgColor.rgb[2:]}" if fill.fgColor.rgb.startswith('FF') else f"#{fill.fgColor.rgb}"

                        if font and font.color and font.color.type == 'rgb' and font.color.rgb:
                            font_color = f"#{font.color.rgb[2:]}" if font.color.rgb.startswith('FF') else f"#{font.color.rgb}"

                        cells.append({
                            "value": cell_value,
                            "bg_color": bg_color,
                            "font_color": font_color,
                        })
                    else:
                        if sval in ['nan', 'none', '<na>', '']:
                            cells.append('')
                        else:
                            cells.append(str(val))

                all_cells.append(cells)

        tc_summary_data: list[list[str]] = []
        if 'TC Summary' in wb.sheetnames:
            ws_tc = wb['TC Summary']
            for row_tc in ws_tc.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row_tc):
                    tc_summary_data.append([str(cell) if cell is not None else '' for cell in row_tc])

        detailed_tc_headers: list[str] = []
        detailed_tc_data: list[list[str]] = []
        if 'Detailed TC Summary' in wb.sheetnames:
            ws_dtc = wb['Detailed TC Summary']
            all_headers: list[str] = []
            wanted_headers = [
                'TC No.', 'TC Name', 'Status', 'Start Date', 'End Date',
                'Total Steps', 'Planned Steps', 'Actual Steps Passed',
                'Planned %', 'Actual %',
            ]

            def norm(h: str) -> str:
                return (
                    str(h).strip().lower().replace(' ', '').replace('%', 'percent')
                )

            header_indices: list[int | None] = []
            for idx2, row_dtc in enumerate(ws_dtc.iter_rows(values_only=True)):
                if idx2 == 2:
                    all_headers = [str(cell).strip() if cell is not None else '' for cell in row_dtc]
                    normed = [norm(h) for h in all_headers]
                    detailed_tc_headers = [wh for wh in wanted_headers]
                    header_indices = []
                    for wh in wanted_headers:
                        nwh = norm(wh)
                        found = None
                        for j, nh in enumerate(normed):
                            if nwh == nh:
                                found = j
                                break
                        header_indices.append(found)
                elif idx2 > 2 and all_headers and detailed_tc_headers:
                    filtered_row: list[str] = []
                    for col_idx, wh in zip(header_indices, wanted_headers):
                        val = row_dtc[col_idx] if col_idx is not None and col_idx < len(row_dtc) else None
                        wh_norm = wh.strip().lower()
                        if wh_norm in ['planned %', 'actual %']:
                            try:
                                if val is not None and str(val).strip() != '' and str(val).strip().lower() not in ['nan', 'none', '<na>']:
                                    percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                    filtered_row.append(f"{percent_val:.2f}%")
                                else:
                                    filtered_row.append('')
                            except Exception:
                                filtered_row.append('')
                        else:
                            filtered_row.append(str(val) if val is not None else '')

                    if any(cell for cell in filtered_row):
                        detailed_tc_data.append(filtered_row)

        open_defects_data: list[list[str]] = []
        if 'Open Defects' in wb.sheetnames:
            ws_od = wb['Open Defects']
            for row_od in ws_od.iter_rows(values_only=True):
                open_defects_data.append([str(cell) if cell is not None else '' for cell in row_od])

        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute(
                f'SELECT columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data FROM {table_name} WHERE filename = ?',
                (filename,),
            )
            existing = cur.fetchone()
            new_data = [
                columns_json,
                json.dumps(all_cells),
                json.dumps(tc_summary_data),
                json.dumps(open_defects_data),
                json.dumps(detailed_tc_headers),
                json.dumps(detailed_tc_data),
            ]

            if existing is None:
                cur.execute(
                    f'INSERT INTO {table_name} (filename, project_name, columns, cells, tc_summary, open_defects, detailed_tc_headers, detailed_tc_data) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                    (
                        filename,
                        project_name,
                        columns_json,
                        json.dumps(all_cells),
                        json.dumps(tc_summary_data),
                        json.dumps(open_defects_data),
                        json.dumps(detailed_tc_headers),
                        json.dumps(detailed_tc_data),
                    ),
                )
                imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
            else:
                if list(existing) != new_data:
                    cur.execute(
                        f'UPDATE {table_name} SET columns=?, cells=?, tc_summary=?, open_defects=?, detailed_tc_headers=?, detailed_tc_data=?, project_name=? WHERE filename=?',
                        (
                            columns_json,
                            json.dumps(all_cells),
                            json.dumps(tc_summary_data),
                            json.dumps(open_defects_data),
                            json.dumps(detailed_tc_headers),
                            json.dumps(detailed_tc_data),
                            project_name,
                            filename,
                        ),
                    )
                    imported_counts[table_name] = imported_counts.get(table_name, 0) + 1
    except Exception:
        traceback.print_exc()


def _import_from_local_folder(folder_path: str | None, tags_filter: set[str] | None = None) -> tuple[str, int]:
    """Import latest matching Excel files from a local folder (e.g., Box Drive-synced path)."""
    if not folder_path:
        return ('Missing folder path. Set BOX_LOCAL_PATH or provide path as query param.', 400)

    if not os.path.isdir(folder_path):
        return (f'Path not found or not a directory: {folder_path}', 404)

    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for fname in files:
            if fname.lower().endswith(('.xlsx', '.xls')):
                full = os.path.join(root, fname)
                excel_files.append({'name': fname, 'path': full})

    latest_file_map: dict[tuple[str, str, str], tuple[tuple[int, int], dict]] = {}
    for f in excel_files:
        filename = f.get('name', '')
        parsed = _parse_dashboard_filename(filename)
        if not parsed:
            continue
        pid, project_name, tag, mm, dd = parsed
        if tags_filter and tag not in tags_filter:
            continue
        key = (pid, project_name, tag)
        if key not in latest_file_map or (mm, dd) > latest_file_map[key][0]:
            latest_file_map[key] = ((mm, dd), f)

    latest_files = [v[1] for v in latest_file_map.values()]

    imported_counts: dict[str, int] = {
        'dashboard_data': 0,
        'dashboard_data_pvt': 0,
        'dashboard_data_billing': 0,
    }

    for file_info in latest_files:
        path = file_info.get('path')
        filename = file_info.get('name')
        if not path or not filename:
            continue
        try:
            with open(path, 'rb') as fh:
                content = fh.read()
        except Exception:
            traceback.print_exc()
            continue

        _process_excel_bytes(filename, content, tags_filter, imported_counts)

    total = sum(imported_counts.values())

    dedup_deleted: dict[str, int] = {}
    if tags_filter:
        for tag in tags_filter:
            table_name = _table_for_tag(tag)
            try:
                dedup_deleted[table_name] = _deduplicate_table_latest_by_pid_project(table_name)
            except Exception:
                traceback.print_exc()

    dedup_msg = ''
    if dedup_deleted:
        parts = [f"{tbl}:{cnt}" for tbl, cnt in dedup_deleted.items() if cnt]
        if parts:
            dedup_msg = f" Dedup deleted {', '.join(parts)} old rows."

    if tags_filter:
        tags_str = ','.join(sorted(tags_filter))
        return (f"Imported {total} files for tags [{tags_str}]. "
                f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}", 200)

    return (f"Imported {total} files. "
            f"ETE={imported_counts.get('dashboard_data',0)}, PVT={imported_counts.get('dashboard_data_pvt',0)}, Billing={imported_counts.get('dashboard_data_billing',0)}.{dedup_msg}", 200)


@app.route('/admin/import_local', methods=['POST'])
def import_from_local():
    # Try query param 'path' first, then env var BOX_LOCAL_PATH, then WATCH_FOLDER
    path = request.args.get('path') or os.environ.get('BOX_LOCAL_PATH') or os.environ.get('WATCH_FOLDER')
    # Optional tag filter via ?tags=ETE,PVT
    tags_q = request.args.get('tags')
    tags_filter = set([t.strip() for t in tags_q.split(',')]) if tags_q else None
    try:
        return _import_from_local_folder(path, tags_filter=tags_filter)
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)


class LocalFolderHandler(FileSystemEventHandler):
    def __init__(self, tags_filter: set[str] | None = None, delay: float = 1.0):
        super().__init__()
        self._jobs: dict[str, float] = {}
        self.tags_filter = tags_filter
        self.delay = delay

    def _schedule(self, path: str):
        # debounce multiple quick events for same path
        now = time.time()
        self._jobs[path] = now
        def worker(p: str, scheduled_at: float):
            time.sleep(self.delay)
            # only run if no newer event for this path
            if self._jobs.get(p) == scheduled_at:
                try:
                    with open(p, 'rb') as fh:
                        content = fh.read()
                    filename = os.path.basename(p)
                    imported_counts = {
                        'dashboard_data': 0,
                        'dashboard_data_pvt': 0,
                        'dashboard_data_billing': 0,
                    }
                    _process_excel_bytes(filename, content, self.tags_filter, imported_counts)
                except Exception:
                    traceback.print_exc()
        t = threading.Thread(target=worker, args=(path, now), daemon=True)
        t.start()

    def on_created(self, event):
        if event.is_directory:
            return
        if not event.src_path.lower().endswith(('.xlsx', '.xls')):
            return
        self._schedule(event.src_path)

    def on_modified(self, event):
        if event.is_directory:
            return
        if not event.src_path.lower().endswith(('.xlsx', '.xls')):
            return
        self._schedule(event.src_path)


def start_folder_watcher(path: str | None, tags_filter: set[str] | None = None):
    if not path:
        return
    if not os.path.isdir(path):
        return
    handler = LocalFolderHandler(tags_filter=tags_filter)
    observer = Observer()
    observer.schedule(handler, path, recursive=True)
    observer.daemon = True
    observer.start()


def send_smtp_email(to_address: str, subject: str, html_body: str) -> tuple[bool, str]:
    """Send email via backend SMTP server (corporate relay). Returns (success, message)."""
    smtp_server = os.environ.get('SMTP_SERVER')
    smtp_port = int(os.environ.get('SMTP_PORT', '25'))
    from_email = os.environ.get('SMTP_FROM_EMAIL')
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    use_tls = os.environ.get('SMTP_USE_TLS', 'false').lower() == 'true'
    
    if not (smtp_server and from_email):
        return (False, 'Missing SMTP_SERVER or SMTP_FROM_EMAIL env vars')
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
        
        if use_tls:
            server.starttls()
        
        if smtp_user and smtp_pass:
            server.login(smtp_user, smtp_pass)
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = to_address
        
        msg.attach(MIMEText(html_body, 'html'))
        
        server.sendmail(from_email, to_address, msg.as_string())
        server.quit()
        
        return (True, f'Email sent successfully to {to_address}')
    except Exception as e:
        return (False, f'Email failed: {str(e)}')


def extract_l2_from_excel(file_path: str) -> tuple[list, list, str, str]:
    """Extract L2 data from Excel file without DB update. Returns (columns, rows, filename, project_name)."""
    from openpyxl import load_workbook
    
    try:
        df = pd.read_excel(file_path, sheet_name=EXCEL_SHEET, header=1)
        selected_cols = [col for col in df.columns if not str(col).startswith('Unnamed')]
        
        wb = load_workbook(file_path, data_only=True)
        if EXCEL_SHEET not in wb.sheetnames:
            return ([], [], '', '')
        ws = wb[EXCEL_SHEET]
        
        status_col_idx = None
        for idx, col in enumerate(selected_cols):
            if str(col).strip().lower() == 'overall status':
                status_col_idx = idx
                break
        
        all_cells = []
        if len(df) > 0 and selected_cols:
            from openpyxl.utils import get_column_letter
            for row_idx, row in df.iterrows():
                project_id_val = row[selected_cols[0]]
                project_id_str = str(project_id_val).strip().lower()
                if not (pd.notnull(project_id_val) and project_id_str not in ['', 'nan', 'none', '<na>']):
                    continue
                
                cells = []
                for col_idx, col in enumerate(selected_cols):
                    val = row[col]
                    sval = str(val).strip().lower()
                    
                    if col_idx == 0:
                        if isinstance(val, float) and val.is_integer():
                            cells.append(str(int(val)))
                        elif sval in ['nan', 'none', '<na>', '']:
                            cells.append('')
                        else:
                            cells.append(str(val))
                    elif str(col).strip().lower() in ['plan start', 'plan end']:
                        try:
                            if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                date_val = pd.to_datetime(val)
                                if date_val.year == 1970 and date_val.month == 1 and date_val.day == 1:
                                    cells.append('')
                                else:
                                    cells.append(date_val.date().strftime('%m/%d/%Y'))
                            else:
                                cells.append('')
                        except Exception:
                            cells.append('')
                    elif str(col).strip().lower() in ['planned %', 'actual %', 'passed %']:
                        try:
                            if pd.notnull(val) and sval not in ['', 'nan', 'none', '<na>']:
                                percent_val = float(val) * 100 if float(val) <= 1 else float(val)
                                cells.append(f"{percent_val:.2f}%")
                            else:
                                cells.append('')
                        except Exception:
                            cells.append('')
                    elif col_idx == status_col_idx:
                        excel_col = get_column_letter(col_idx + 1)
                        excel_row = row_idx + 3
                        cell_obj = ws[f"{excel_col}{excel_row}"]
                        fill = cell_obj.fill
                        font = cell_obj.font
                        bg_color = None
                        font_color = None
                        
                        cell_value = str(val) if sval not in ['nan', 'none', '<na>', ''] else ''
                        status_val = cell_value.strip().upper()
                        if status_val == 'NOT STARTED':
                            bg_color = None
                        else:
                            if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
                                bg_color = f"#{fill.fgColor.rgb[2:]}" if fill.fgColor.rgb.startswith('FF') else f"#{fill.fgColor.rgb}"
                        
                        if font and font.color and font.color.type == 'rgb' and font.color.rgb:
                            font_color = f"#{font.color.rgb[2:]}" if font.color.rgb.startswith('FF') else f"#{font.color.rgb}"
                        
                        cells.append({
                            "value": cell_value,
                            "bg_color": bg_color,
                            "font_color": font_color,
                        })
                    else:
                        if sval in ['nan', 'none', '<na>', '']:
                            cells.append('')
                        else:
                            cells.append(str(val))
                
                all_cells.append(cells)
        
        # Parse filename to get project name
        filename = os.path.basename(file_path)
        parsed = _parse_dashboard_filename(filename)
        project_name = parsed[1] if parsed else ''
        
        return (selected_cols, all_cells, filename, project_name)
    except Exception:
        traceback.print_exc()
        return ([], [], '', '')


def generate_l2_html_report(box_folder: str) -> tuple[str, str]:
    """Read latest Excel files from Box folder and generate HTML report. Returns (html, summary)."""
    if not os.path.isdir(box_folder):
        return ('', f'Path not found: {box_folder}')
    
    excel_files = []
    for root, dirs, files in os.walk(box_folder):
        for fname in files:
            if fname.lower().endswith(('.xlsx', '.xls')):
                full = os.path.join(root, fname)
                excel_files.append({'name': fname, 'path': full})
    
    # Keep only latest per (PID, project_name, tag)
    latest_file_map = {}
    for f in excel_files:
        filename = f.get('name', '')
        parsed = _parse_dashboard_filename(filename)
        if not parsed:
            continue
        pid, project_name, tag, mm, dd = parsed
        key = (pid, project_name, tag)
        if key not in latest_file_map or (mm, dd) > latest_file_map[key][0]:
            latest_file_map[key] = ((mm, dd), f)
    
    latest_files = [v[1] for v in latest_file_map.values()]
    
    # Extract L2 data from all latest files
    all_data = []
    for file_info in latest_files:
        path = file_info.get('path')
        filename = file_info.get('name')
        if not path or not filename:
            continue
        columns, rows, fname, project_name = extract_l2_from_excel(path)
        if columns and rows:
            all_data.append({
                'filename': fname,
                'project_name': project_name,
                'columns': columns,
                'rows': rows,
            })
    
    if not all_data:
        return ('', 'No valid L2 data found in Box folder')
    
    # Generate HTML similar to dashboard
    html_parts = [
        '<!DOCTYPE html>',
        '<html lang="en">',
        '<head>',
        '<meta charset="UTF-8">',
        '<title>L2 Project Dashboard Report</title>',
        '<style>',
        'body { font-family: "Times New Roman", serif; background: #f7f3ea; padding: 20px; }',
        'h1 { text-align: center; color: #7a5c2e; }',
        'h2 { color: #7a5c2e; margin-top: 30px; }',
        'table { border-collapse: collapse; width: 100%; margin: 20px 0; border: 1.5px solid #e0c9a6; background: #fdf6ee; }',
        'th { background: #f5e9da; color: #7a5c2e; font-weight: bold; border: 1px solid #e0c9a6; padding: 8px; text-align: center; }',
        'td { border: 1px solid #e0c9a6; padding: 8px; text-align: center; }',
        'tr:nth-child(even) { background: #f9f3e7; }',
        'tr:nth-child(odd) { background: #fdf6ee; }',
        'tr:hover { background: #f3e7d1; }',
        '.status-badge { display: inline-block; border-radius: 16px; padding: 6px 18px; min-width: 90px; color: #fff; font-weight: bold; border: 1.5px solid #fff; box-shadow: 0 2px 6px rgba(0,0,0,0.08); }',
        '</style>',
        '</head>',
        '<body>',
        '<h1>L2 Project Dashboard Report</h1>',
    ]
    
    for data in all_data:
        html_parts.append(f"<h2>{data['project_name']} - {data['filename']}</h2>")
        html_parts.append('<table>')
        
        # Headers
        html_parts.append('<thead><tr>')
        for col in data['columns']:
            html_parts.append(f'<th>{col}</th>')
        html_parts.append('</tr></thead>')
        
        # Rows
        html_parts.append('<tbody>')
        for row_cells in data['rows']:
            html_parts.append('<tr>')
            for cell_idx, cell in enumerate(row_cells):
                if isinstance(cell, dict):
                    # Status cell with color
                    bg_color = cell.get('bg_color', '')
                    font_color = cell.get('font_color', '#fff')
                    value = cell.get('value', '')
                    if bg_color:
                        html_parts.append(f'<td><span class="status-badge" style="background-color: {bg_color}; color: {font_color};">{value}</span></td>')
                    else:
                        html_parts.append(f'<td>{value}</td>')
                else:
                    html_parts.append(f'<td>{cell}</td>')
            html_parts.append('</tr>')
        html_parts.append('</tbody>')
        html_parts.append('</table>')
    
    html_parts.extend([
        '</body>',
        '</html>',
    ])
    
    html_body = '\n'.join(html_parts)
    summary = f'Report generated with {len(all_data)} project(s)'
    return (html_body, summary)


@app.route('/admin/send_l2_report', methods=['POST'])
def send_l2_report():
    """Read Box folder, generate L2 report, and email it."""
    box_path = os.environ.get('BOX_LOCAL_PATH') or os.environ.get('WATCH_FOLDER')
    to_email = request.args.get('to') or request.form.get('to')
    
    if not box_path:
        return ('Missing BOX_LOCAL_PATH or WATCH_FOLDER env var', 400)
    if not to_email:
        return ('Missing email address. Provide ?to=email@example.com or form field to=', 400)
    
    try:
        html_report, summary = generate_l2_html_report(box_path)
        if not html_report:
            return (f'Failed to generate report: {summary}', 500)
        
        success, msg = send_smtp_email(to_email, 'L2 Project Dashboard Report', html_report)
        if success:
            return (f'{msg}. {summary}', 200)
        else:
            return (msg, 500)
    except Exception as e:
        traceback.print_exc()
        return (f'Error: {str(e)}', 500)



def _maybe_start_watcher():
    # Start watcher only once. Prefer explicit env var BOX_LOCAL_PATH or WATCH_FOLDER
    path = os.environ.get('BOX_LOCAL_PATH') or os.environ.get('WATCH_FOLDER')
    if not path:
        return
    # Avoid starting multiple times under dev reloader
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not os.environ.get('WERKZEUG_RUN_MAIN'):
        # Start in background thread so it doesn't block request handling
        t = threading.Thread(target=start_folder_watcher, args=(path, None), daemon=True)
        t.start()


# --- ADMIN: Import from GitHub to DB (should be protected in production) ---
@app.route('/admin/import', methods=['POST'])
def import_from_github():
    try:
        return _import_from_github_routed(env_prefix='ETE', tags_filter={'ETE'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)


@app.route('/admin/import_pvt', methods=['POST'])
def import_from_github_pvt():
    try:
        return _import_from_github_routed(env_prefix='PVT', tags_filter={'PVT'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)


@app.route('/admin/import_billing', methods=['POST'])
def import_from_github_billing():
    try:
        return _import_from_github_routed(env_prefix='BILLING', tags_filter={'ETE_Billing'})
    except Exception as e:
        traceback.print_exc()
        return (f'Import failed: {e}', 500)
    
    # The following code block was incorrectly indented and unreachable; it should be part of a function, such as download_and_read_columns, or moved to the correct location.
    # Please move this logic into the appropriate function if needed.
    # For now, it is commented out to avoid syntax errors.
    
    #     # Read Detailed TC Summary sheet headers and data (row 3, index 2)
    #     detailed_tc_headers = None
    #     detailed_tc_data = []
    #     if 'Detailed TC Summary' in wb.sheetnames:
    #         ws_dtc = wb['Detailed TC Summary']
    #         all_headers = []
    #         for idx, row in enumerate(ws_dtc.iter_rows(values_only=True)):
    #             if idx == 2:
    #                 all_headers = [str(cell) if cell is not None else '' for cell in row]
    #                 wanted_headers = [
    #                     'TC No.', 'TC Name', 'Status', 'Start Date', 'End Date',
    #                     'Total Steps', 'Planned Steps', 'Actual Steps Passed'
    #                 ]
    #                 filtered_headers = [h for h in all_headers if h in wanted_headers]
    #                 detailed_tc_headers = filtered_headers
    #             elif idx > 2 and all_headers:
    #                 # Only add rows if headers are set
    #                 filtered_row = []
    #                 for h in detailed_tc_headers:
    #                     try:
    #                         col_idx = all_headers.index(h)
    #                         val = row[col_idx]
    #                         # If Start Date or End Date, format as date only
    #                         if h in ['Start Date', 'End Date'] and val:
    #                             from datetime import datetime
    #                             if isinstance(val, datetime):
    #                                 filtered_row.append(val.strftime('%m/%d/%Y'))
    #                             else:
    #                                 sval = str(val)
    #                                 # Try to parse date from string
    #                                 try:
    #                                     dt = pd.to_datetime(sval)
    #                                     filtered_row.append(dt.strftime('%m/%d/%Y'))
    #                                 except Exception:
    #                                     filtered_row.append(sval.split()[0])
    #                         else:
    #                             filtered_row.append(str(val) if val is not None else '')
    #                     except Exception:
    #                         filtered_row.append('')
    #                 if any(cell for cell in filtered_row):
    #                     detailed_tc_data.append(filtered_row)
    #     # ...existing code...
    #     # Read Open Defects sheet if present
    #     open_defects_data = None
    #     open_defects_header = None
    #     if 'Open Defects' in wb.sheetnames:
    #         ws_od = wb['Open Defects']
    #         open_defects_data = []
    #         for idx, row_od in enumerate(ws_od.iter_rows(values_only=True)):
    #             if idx == 0:
    #                 open_defects_header = [str(cell) if cell is not None else '' for cell in row_od]
    #                 continue
    #             processed_row = []
    #             for col_idx, cell in enumerate(row_od):
    #                 if cell is None:
    #                     processed_row.append('')
    #                 else:
    #                     # Check if this is the 'Defect Opened' column
    #                     is_defect_opened = open_defects_header and col_idx < len(open_defects_header) and open_defects_header[col_idx].strip().lower() == 'defect opened'
    #                     if is_defect_opened:
    #                         from datetime import datetime
    #                         if isinstance(cell, str):
    #                             try:
    #                                 dt = datetime.strptime(cell, '%Y-%m-%d %H:%M:%S')
    #                                 processed_row.append(dt.strftime('%Y-%m-%d'))
    #                             except Exception:
    #                                 # Try parsing as just date
    #                                 try:
    #                                     dt = datetime.strptime(cell, '%Y-%m-%d')
    #                                     processed_row.append(dt.strftime('%Y-%m-%d'))
    #                                 except Exception:
    #                                     processed_row.append(str(cell))
    #                         elif isinstance(cell, (datetime,)):
    #                             processed_row.append(cell.strftime('%Y-%m-%d'))
    #                         else:
    #                             processed_row.append(str(cell))
    #                     else:
    #                         val = str(cell)
    #                         val = val.replace('\r\n', '___LINEBREAK___').replace('\n', '___LINEBREAK___').replace('\r', '___LINEBREAK___')
    #                         processed_row.append(val)
    #             open_defects_data.append(processed_row)
    #     # Attach header as first row if present
    #     if open_defects_data is not None and open_defects_header is not None:
    #         open_defects_data = [open_defects_header] + open_defects_data
    #     rows = []
    #     if len(df) > 0:
    #         for idx, row in df.iterrows():
    #             project_id_val = row[selected_cols[0]]
    #             project_id_str = str(project_id_val).strip().lower()
    #             if not (pd.notnull(project_id_val) and project_id_str != '' and project_id_str != 'nan' and project_id_str != 'none' and project_id_str != '<na>'):
    #                 continue
    #             cells = []
    #             for col_idx, col in enumerate(selected_cols):
    #                 val = row[col]
    #                 sval = str(val).strip().lower()
    #                 if col_idx == 0:
    #                     if isinstance(val, float) and val.is_integer():
    #                         cells.append(str(int(val)))
    #                     elif sval == 'nan' or sval == 'none' or sval == '<na>' or sval == '':
    #                         cells.append('')
    #                     else:
    #                         cells.append(str(val))
    #                 elif str(col).strip().lower() in ['plan start', 'plan end']:
    #                     try:
    #                         if pd.notnull(val) and sval != '' and sval != 'nan' and sval != 'none' and sval != '<na>':
    #                             date_val = pd.to_datetime(val)
    #                             if date_val.year == 1970 and date_val.month == 1 and date_val.day == 1:
    #                                 cells.append('')
    #                             else:
    #                                 cells.append(date_val.date().strftime('%m/%d/%Y'))
    #                         else:
    #                             cells.append('')
    #                     except Exception:
    #                         cells.append('')
    #                 elif str(col).strip().lower() in ['planned %', 'actual %', 'passed %']:
    #                     try:
    #                         if pd.notnull(val) and sval != '' and sval != 'nan' and sval != 'none' and sval != '<na>':
    #                             percent_val = float(val) * 100 if float(val) <= 1 else float(val)
    #                             cells.append(f"{percent_val:.2f}%")
    #                         else:
    #                             cells.append('')
    #                     except Exception:
    #                         cells.append('')
    #                 elif col_idx == status_col_idx:
    #                     excel_col = get_column_letter(col_idx + 1)
    #                     excel_row = idx + 3  # header=1, so first data row is Excel row 3
    #                     cell_obj = ws[f"{excel_col}{excel_row}"]
    #                     fill = cell_obj.fill
    #                     font = cell_obj.font
    #                     bg_color = None
    #                     font_color = None
    #                     # Remove blue background for 'NOT STARTED' status
    #                     cell_value = str(val) if sval not in ['nan', 'none', '<na>', ''] else ''
    #                     status_val = cell_value.strip().upper()
    #                     if status_val == 'NOT STARTED':
    #                         bg_color = None  # Always no background for NOT STARTED
    #                     else:
    #                         if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
    #                             bg_color = f"#{fill.fgColor.rgb[2:]}" if fill.fgColor.rgb.startswith('FF') else f"#{fill.fgColor.rgb}"
    #                     if font and font.color and font.color.type == 'rgb' and font.color.rgb:
    #                         font_color = f"#{font.color.rgb[2:]}" if font.color.rgb.startswith('FF') else f"#{font.color.rgb}"
    #                     cells.append({
    #                         "value": cell_value,
    #                         "bg_color": bg_color,
    #                         "font_color": font_color
    #                     })
    #                 else:
    #                     if sval == 'nan' or sval == 'none' or sval == '<na>' or sval == '':
    #                         cells.append('')
    #                     else:
    #                         cells.append(str(val))
    #             # Attach TC Summary, Open Defects, and Detailed TC Summary to each row (same for all rows in file)
    #             rows.append({
    #                 "cells": cells,
    #                 "tc_summary": tc_summary_data if tc_summary_data else [],
    #                 "open_defects": open_defects_data if open_defects_data else [],
    #                 "detailed_tc_headers": detailed_tc_headers if detailed_tc_headers else [],
    #                 "detailed_tc_data": detailed_tc_data if detailed_tc_data else []
    #             })
    #     return {"rows": rows, "columns": [str(c) for c in selected_cols], "error": None, "status_col_idx": status_col_idx}
    # # except Exception as e:
    # #     return {"rows": [], "columns": [], "error": str(e), "status_col_idx": None}

# Main dashboard route
@app.route('/')
def main_dashboard():
    html_path = os.path.join(os.path.dirname(__file__), 'main_dashboard.html')
    return send_file(html_path)


# ETE Project Dashboard route (now fetches from DB)
@app.route('/ete')
def ete_dashboard():
    # Read dashboard-ready rows from dashboard_data table
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT columns, cells, project_name FROM dashboard_data')
    records = cursor.fetchall()
    conn.close()
    if not records:
        return render_template_string(dashboard_html, columns=[], rows=[], status_col_idx=None, errors=["No data in DB. Please import from GitHub."])
    # Reconstruct columns and rows
    all_rows = []
    columns = None
    dedup_dict = {}
    for rec in records:
        rec_columns = json.loads(rec[0]) if rec[0] else None
        all_cells = json.loads(rec[1])
        project_name = rec[2] if len(rec) > 2 else ''
        if columns is None and rec_columns:
            columns = rec_columns
            # Insert Project Name as the second column if not present
            if 'Project Name' not in columns:
                columns = columns[:1] + ['Project Name'] + columns[1:]
        for cells in all_cells:
            # Insert project_name as the second cell if not present
            if 'Project Name' in columns and (len(cells) < len(columns)):
                cells = cells[:1] + [project_name] + cells[1:]
            # Deduplication key: (Project ID, Project Name)
            try:
                pid_idx = columns.index('Project ID')
                pname_idx = columns.index('Project Name')
                pid = str(cells[pid_idx]).strip()
                pname = str(cells[pname_idx]).strip()
                key = (pid, pname)
                dedup_dict[key] = {
                    "cells": cells
                }
            except Exception:
                # fallback: just append
                all_rows.append({
                    "cells": cells
                })
    # Use deduplicated rows if any
    if dedup_dict:
        all_rows = list(dedup_dict.values())
    # Find status_col_idx
    status_col_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'overall status':
            status_col_idx = idx
            break
    # Default sort by Plan Start descending if column exists
    plan_start_idx = None
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == 'plan start':
            plan_start_idx = idx
            break
    if plan_start_idx is not None:
        from datetime import datetime
        def parse_date(val):
            try:
                return datetime.strptime(val, '%m/%d/%Y')
            except Exception:
                return datetime.min
        all_rows = sorted(
            all_rows,
            key=lambda row: parse_date(row['cells'][plan_start_idx]) if len(row['cells']) > plan_start_idx else datetime.min,
            reverse=True
        )
    return render_template_string(
        dashboard_html,
        dashboard_title='ETE Project Dashboard - L2 Report',
        columns=columns if columns else [],
        rows=all_rows,
        status_col_idx=status_col_idx,
        errors=[],
    )

# --- Email Report Routes (for testing before automation) ---
@app.route('/email/test-format')
def email_test_format():
    """Display test email format HTML file."""
    html_path = os.path.join(os.path.dirname(__file__), 'test_email_format.html')
    if os.path.exists(html_path):
        return send_file(html_path)
    else:
        return "Test email format file not found", 404


@app.route('/email/generate-report')
def email_generate_report():
    """Generate and display HTML report (preview)."""
    try:
        from report_generator import get_report_generator
        generator = get_report_generator()
        html = generator.generate_email_html(
            subject=f"L2 Project Dashboard Report - {datetime.now().strftime('%Y-%m-%d')}"
        )
        return html
    except Exception as e:
        return f"<h2>Error generating report:</h2><p>{str(e)}</p>", 500


@app.route('/email/send-test', methods=['GET', 'POST'])
def email_send_test():
    """Send test email (requires recipient email in query param or form)."""
    if request.method == 'GET':
        # Show form
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body { font-family: 'Times New Roman', Times, serif; margin: 40px; }
                .form-container { max-width: 600px; margin: 0 auto; padding: 20px; 
                                 border: 1px solid #e0c9a6; border-radius: 8px; 
                                 background: #fdf6ee; }
                h2 { color: #7a5c2e; }
                label { display: block; margin-top: 15px; color: #7a5c2e; }
                input { width: 100%; padding: 8px; border: 1px solid #e0c9a6; 
                       border-radius: 4px; box-sizing: border-box; }
                button { margin-top: 20px; padding: 10px 20px; 
                        background: #7a5c2e; color: white; border: none; 
                        border-radius: 4px; cursor: pointer; }
                button:hover { background: #5a3d13; }
                .info { background: #e8f5e9; padding: 10px; margin-bottom: 20px; 
                       border-radius: 4px; color: #2e7d32; }
            </style>
        </head>
        <body>
            <div class="form-container">
                <h2>Send Test Email</h2>
                <div class="info">
                    <strong>Info:</strong> This will send a test report email to verify your SMTP configuration.
                </div>
                <form method="POST">
                    <label for="email">Recipient Email Address:</label>
                    <input type="email" id="email" name="email" required placeholder="your.email@company.com">
                    
                    <label for="cc">CC Email (optional):</label>
                    <input type="email" id="cc" name="cc" placeholder="optional.cc@company.com">
                    
                    <button type="submit">Send Test Report</button>
                </form>
            </div>
        </body>
        </html>
        """
    else:
        # Send email
        try:
            recipient_email = request.form.get('email', '').strip()
            cc_email = request.form.get('cc', '').strip()
            
            if not recipient_email:
                return "Recipient email is required", 400
            
            from email_utils import EmailSender
            from report_generator import get_report_generator
            
            # Generate report
            generator = get_report_generator()
            subject = f"L2 Project Dashboard Test Report - {datetime.now().strftime('%Y-%m-%d')}"
            html_content = generator.generate_email_html(subject=subject)
            
            # Send email
            sender = EmailSender()
            cc_list = [cc_email] if cc_email else None
            
            success = sender.send_email(
                to_emails=[recipient_email],
                subject=subject,
                html_content=html_content,
                cc_emails=cc_list,
                text_content=f"L2 Project Dashboard Report\n\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            
            if success:
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <style>
                        body {{ font-family: 'Times New Roman', Times, serif; margin: 40px; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; 
                                     border: 1px solid #4caf50; border-radius: 8px; 
                                     background: #e8f5e9; }}
                        h2 {{ color: #2e7d32; }}
                        p {{ color: #333; }}
                        a {{ color: #2e7d32; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h2>✓ Test Email Sent Successfully!</h2>
                        <p><strong>Sent to:</strong> {recipient_email}</p>
                        {f'<p><strong>CC:</strong> {cc_email}</p>' if cc_email else ''}
                        <p>Check your email inbox for the test report.</p>
                        <p><a href="/email/send-test">Send another test</a> | <a href="/">Back to Dashboard</a></p>
                    </div>
                </body>
                </html>
                """
            else:
                return f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <style>
                        body {{ font-family: 'Times New Roman', Times, serif; margin: 40px; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; 
                                     border: 1px solid #f44336; border-radius: 8px; 
                                     background: #ffebee; }}
                        h2 {{ color: #d32f2f; }}
                        p {{ color: #333; }}
                        a {{ color: #d32f2f; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h2>✗ Failed to Send Test Email</h2>
                        <p>Check your SMTP configuration in the .env file.</p>
                        <p><a href="/email/send-test">Try again</a> | <a href="/">Back to Dashboard</a></p>
                    </div>
                </body>
                </html>
                """
                
        except Exception as e:
            return f"<h2>Error:</h2><p>{str(e)}</p><p><a href='/email/send-test'>Back</a></p>", 500


@app.route('/scheduler/info')
def scheduler_info():
    """Display scheduler information."""
    try:
        from report_scheduler import get_scheduler
        scheduler = get_scheduler()
        jobs = scheduler.get_jobs()
        
        jobs_html = ""
        if jobs:
            jobs_html = "<ul>"
            for job in jobs:
                jobs_html += f"""
                <li>
                    <strong>{job.name}</strong> (ID: {job.id})<br>
                    Trigger: {job.trigger}<br>
                    Next run: {job.next_run_time}
                </li>
                """
            jobs_html += "</ul>"
        else:
            jobs_html = "<p><em>No jobs scheduled yet.</em></p>"
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: 'Times New Roman', Times, serif; margin: 40px; }}
                .container {{ max-width: 700px; margin: 0 auto; padding: 20px; 
                             border: 1px solid #e0c9a6; border-radius: 8px; 
                             background: #fdf6ee; }}
                h2 {{ color: #7a5c2e; }}
                ul {{ margin: 15px 0; }}
                li {{ margin: 10px 0; padding: 10px; background: white; 
                    border-left: 3px solid #7a5c2e; border-radius: 4px; }}
                a {{ color: #1976d2; text-decoration: none; margin-right: 15px; }}
                a:hover {{ text-decoration: underline; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Scheduler Information</h2>
                <p><strong>Status:</strong> {'Running' if scheduler.scheduler and scheduler.scheduler.running else 'Not Running'}</p>
                <p><strong>Recipients:</strong> {', '.join(scheduler.recipients)}</p>
                <h3>Scheduled Jobs:</h3>
                {jobs_html}
                <p>
                    <a href="/">Back to Dashboard</a> | 
                    <a href="/email/send-test">Send Test Email</a>
                </p>
            </div>
        </body>
        </html>
        """
    except Exception as e:
        return f"<h2>Error:</h2><p>{str(e)}</p>", 500


if __name__ == '__main__':
    try:
        print('Starting Flask app on http://localhost:5000 ...')
        # start folder watcher if configured
        try:
            _maybe_start_watcher()
        except Exception:
            traceback.print_exc()
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
    except KeyboardInterrupt:
        print('\nServer stopped by user (Ctrl+C)')


